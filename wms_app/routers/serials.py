from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session
from typing import List
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from wms_app import models, schemas
from wms_app.database import get_db
from wms_app.routers.auth import require_permission
from wms_app.services.serial_service import SerialService
from wms_app.services.logging_service import LoggingService
from wms_app.models.logs import OperationType, OperationCategory, OperationStatus

# Import templates in modo lazy per evitare import circolari
def get_templates():
    from wms_app.main import templates
    return templates

router = APIRouter(
    prefix="/serials",
    tags=["serials"],
)

# --- Viste HTML ---

@router.get("/manage", response_class=HTMLResponse)
async def get_serials_management_page(request: Request, db: Session = Depends(get_db)): 
    """Pagina di gestione seriali prodotto"""
    serial_service = SerialService(db)
    orders_with_serials = serial_service.get_orders_with_serials()
    
    return get_templates().TemplateResponse("serials.html", {
        "request": request,
        "orders_with_serials": orders_with_serials,
        "active_page": "serials"
    })

# --- API Endpoints ---

@router.post("/upload", response_model=schemas.serials.SerialUploadResult)
async def upload_serials_file(
    file: UploadFile = File(...), 
    uploaded_by: str = "system",
    db: Session = Depends(get_db)
):
    """
    Upload e parsing del file seriali
    Formato atteso: ORDINE,EAN,SERIALE,EAN,SERIALE,...
    """
    # Verifica formato file
    if not file.filename.endswith(('.txt', '.csv')):
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usare .txt o .csv")
    
    try:
        # Leggi contenuto file
        content = await file.read()
        file_content = content.decode('utf-8')
        
        # Processa con SerialService
        serial_service = SerialService(db)
        result = serial_service.parse_serial_file(file_content, uploaded_by, file.filename)
        
        if not result.success:
            # Non raise HTTPException, restituisci il result per mostrare gli errori
            return result
        
        return result
        
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Errore codifica file. Usare codifica UTF-8.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante il caricamento: {str(e)}")

@router.get("/orders", response_model=List[schemas.serials.OrderSerialsView])
def get_orders_with_serials(db: Session = Depends(get_db)):
    """Lista ordini con seriali caricati"""
    serial_service = SerialService(db)
    return serial_service.get_orders_with_serials()

@router.get("/orders/{order_number}", response_model=schemas.serials.OrderSerialsView)
def get_order_serials(order_number: str, db: Session = Depends(get_db)):
    """Dettaglio seriali per un ordine specifico"""
    serial_service = SerialService(db)
    return serial_service.get_order_serials_view(order_number)

@router.get("/orders/{order_number}/validate", response_model=schemas.serials.SerialValidationSummary)
def validate_order_serials(order_number: str, db: Session = Depends(get_db)):
    """Validazione seriali di un ordine"""
    serial_service = SerialService(db)
    return serial_service.validate_serials_for_order(order_number)

@router.get("/orders/{order_number}/pdf")
async def generate_serials_pdf(order_number: str, db: Session = Depends(get_db)):
    """
    Genera PDF con seriali per un ordine
    """
    serial_service = SerialService(db)
    order_view = serial_service.get_order_serials_view(order_number)
    
    if not order_view.order_exists:
        raise HTTPException(status_code=404, detail=f"Ordine {order_number} non trovato")
    
    if not order_view.found_serials:
        raise HTTPException(status_code=404, detail=f"Nessun seriale trovato per ordine {order_number}")
    
    # Genera PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Crea stili personalizzati
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Centrato
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20
    )
    
    story = []
    
    # Titolo
    story.append(Paragraph(f"SERIALI PRODOTTO - ORDINE {order_number}", title_style))
    story.append(Spacer(1, 12))
    
    # Data generazione
    current_date = datetime.now().strftime("%d/%m/%Y alle %H:%M")
    story.append(Paragraph(f"Documento generato il {current_date}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Tabella seriali per prodotto
    story.append(Paragraph("DETTAGLIO SERIALI PER PRODOTTO", subtitle_style))
    
    for sku in sorted(order_view.found_serials.keys()):
        serials = order_view.found_serials[sku]
        expected_qty = order_view.expected_products.get(sku, 0)
        found_qty = len(serials)
        
        # Sottotitolo prodotto
        product_title = f"SKU: {sku} (Attesi: {expected_qty}, Trovati: {found_qty})"
        story.append(Paragraph(product_title, styles['Heading3']))
        
        # Tabella seriali per questo prodotto
        serial_data = [["#", "Numero Seriale"]]
        for i, serial in enumerate(serials, 1):
            serial_data.append([str(i), serial])
        
        serial_table = Table(serial_data, colWidths=[0.5*inch, 4*inch])
        serial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(serial_table)
        story.append(Spacer(1, 15))
    
    # Costruisci PDF
    doc.build(story)
    buffer.seek(0)
    
    # Ritorna response
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=seriali_ordine_{order_number}.pdf"}
    )

@router.get("/orders/{order_number}/csv")
async def generate_serials_csv(order_number: str, db: Session = Depends(get_db)):
    """
    Genera CSV con seriali per un ordine
    """
    serial_service = SerialService(db)
    order_view = serial_service.get_order_serials_view(order_number)
    
    if not order_view.order_exists:
        raise HTTPException(status_code=404, detail=f"Ordine {order_number} non trovato")
    
    if not order_view.found_serials:
        raise HTTPException(status_code=404, detail=f"Nessun seriale trovato per ordine {order_number}")
    
    # Genera CSV content
    csv_content = f"# SERIALI PRODOTTO - ORDINE {order_number}\n"
    csv_content += f"# Documento generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}\n"
    csv_content += "SKU,Numero_Seriale,Posizione\n"
    
    # Dati seriali per prodotto
    for sku in sorted(order_view.found_serials.keys()):
        serials = order_view.found_serials[sku]
        
        for i, serial in enumerate(serials, 1):
            csv_content += f'"{sku}","{serial}",{i}\n'
    
    # Ritorna response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=seriali_ordine_{order_number}.csv"}
    )

@router.get("/orders/{order_number}/excel")
async def generate_serials_excel(order_number: str, db: Session = Depends(get_db)):
    """
    Genera Excel con seriali per un ordine usando il template
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export non disponibile. Installare openpyxl.")
    
    serial_service = SerialService(db)
    order_view = serial_service.get_order_serials_view(order_number)
    
    if not order_view.order_exists:
        raise HTTPException(status_code=404, detail=f"Ordine {order_number} non trovato")
    
    if not order_view.found_serials:
        raise HTTPException(status_code=404, detail=f"Nessun seriale trovato per ordine {order_number}")
    
    # Carica il template Excel
    try:
        from openpyxl import load_workbook
        template_path = "/mnt/c/WMS_EPM/Esempio file excel/Template Seriali.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Rinomina il sheet con numero ordine
        ws.title = f"Seriali ordine n {order_number}"
        
    except Exception as e:
        # Fallback: crea workbook manualmente con stili del template
        wb = Workbook()
        ws = wb.active
        ws.title = f"Seriali ordine n {order_number}"
        
        # Stili del template - con controllo import
        if EXCEL_AVAILABLE:
            from openpyxl.styles import Font, Alignment
            header_font = Font(name="Aptos Narrow", bold=True, size=11)
            header_alignment = Alignment(horizontal='center')
        else:
            header_font = None
            header_alignment = None
        
        # Crea header come nel template
        headers = ['Ordine', 'Codice Prodotto', 'Seriale prodotto', 'Data']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if header_font:
                cell.font = header_font
            if header_alignment:
                cell.alignment = header_alignment
    
    # Ottieni i seriali dal database per avere le date
    serials_from_db = db.query(models.serials.ProductSerial).filter(
        models.serials.ProductSerial.order_number == order_number
    ).order_by(
        models.serials.ProductSerial.product_sku,
        models.serials.ProductSerial.serial_number
    ).all()
    
    if not serials_from_db:
        raise HTTPException(status_code=404, detail=f"Nessun seriale trovato per ordine {order_number}")
    
    # Raggruppa i seriali per SKU per calcolare i range di merge
    from collections import OrderedDict
    serials_by_sku = OrderedDict()
    for serial in serials_from_db:
        if serial.product_sku not in serials_by_sku:
            serials_by_sku[serial.product_sku] = []
        serials_by_sku[serial.product_sku].append(serial)
    
    # Popola i dati partendo dalla riga 2
    current_row = 2
    first_row = 2
    
    # Data comune (prendi dalla prima entry) - solo data senza orario
    first_serial = serials_from_db[0]
    if first_serial.uploaded_at:
        common_date = first_serial.uploaded_at.strftime("%d/%m/%Y")
    elif first_serial.created_at:
        common_date = first_serial.created_at.strftime("%d/%m/%Y")
    else:
        common_date = "N/A"
    
    # Processa ogni SKU
    for sku, sku_serials in serials_by_sku.items():
        sku_start_row = current_row
        
        # Scrivi i seriali per questo SKU
        for serial in sku_serials:
            ws.cell(row=current_row, column=1, value=serial.order_number)  # Ordine (sarà unito dopo)
            ws.cell(row=current_row, column=2, value=serial.product_sku)   # Codice Prodotto (sarà unito dopo)
            ws.cell(row=current_row, column=3, value=serial.serial_number) # Seriale prodotto (non unito)
            ws.cell(row=current_row, column=4, value=common_date)          # Data (sarà unita dopo)
            current_row += 1
        
        # Unisci celle per questo SKU (colonna B - Codice Prodotto)
        if len(sku_serials) > 1:
            sku_end_row = current_row - 1
            ws.merge_cells(f'B{sku_start_row}:B{sku_end_row}')
    
    # Unisci celle per tutto l'ordine
    last_row = current_row - 1
    if last_row > first_row:
        # Colonna A - Numero Ordine (tutto l'ordine)
        ws.merge_cells(f'A{first_row}:A{last_row}')
        
        # Colonna D - Data (tutto l'ordine)  
        ws.merge_cells(f'D{first_row}:D{last_row}')
    
    # Allineamento verticale centrato per le celle unite
    from openpyxl.styles import Alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Applica allineamento alle celle della colonna A (Ordine) e D (Data)
    for row_num in range(first_row, last_row + 1):
        ws.cell(row=row_num, column=1).alignment = center_alignment  # Ordine
        ws.cell(row=row_num, column=4).alignment = center_alignment  # Data
    
    # Applica allineamento alle celle unite dei codici prodotto (colonna B)
    current_row = first_row
    for sku, sku_serials in serials_by_sku.items():
        for i, serial in enumerate(sku_serials):
            if i == 0:  # Solo la prima cella di ogni gruppo SKU
                ws.cell(row=current_row, column=2).alignment = center_alignment
            current_row += 1
    
    # Applica bordi e spaziatura alle celle
    from openpyxl.styles import Border, Side
    
    # Definisci stili dei bordi
    thin_border = Side(border_style="thin", color="000000")
    thick_border = Side(border_style="thick", color="000000")
    
    # Bordo interno normale
    normal_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )
    
    # Bordi esterni spessi
    top_border = Border(top=thick_border, left=thin_border, right=thin_border, bottom=thin_border)
    bottom_border = Border(bottom=thick_border, left=thin_border, right=thin_border, top=thin_border)
    left_border = Border(left=thick_border, right=thin_border, top=thin_border, bottom=thin_border)
    right_border = Border(right=thick_border, left=thin_border, top=thin_border, bottom=thin_border)
    
    # Bordi angoli spessi
    top_left_border = Border(top=thick_border, left=thick_border, right=thin_border, bottom=thin_border)
    top_right_border = Border(top=thick_border, right=thick_border, left=thin_border, bottom=thin_border)
    bottom_left_border = Border(bottom=thick_border, left=thick_border, right=thin_border, top=thin_border)
    bottom_right_border = Border(bottom=thick_border, right=thick_border, left=thin_border, top=thin_border)
    
    # Applica bordi e spaziatura a tutte le celle (header + dati)
    for row_num in range(1, last_row + 1):
        for col_num in range(1, 5):  # Colonne A, B, C, D
            cell = ws.cell(row=row_num, column=col_num)
            
            # Spaziatura interna nelle celle (padding)
            if not cell.alignment:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # Mantieni l'allineamento esistente ma assicura il padding
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or 'center',
                    vertical=cell.alignment.vertical or 'center'
                )
            
            # Determina il tipo di bordo in base alla posizione
            if row_num == 1 and col_num == 1:  # Top-left corner
                cell.border = top_left_border
            elif row_num == 1 and col_num == 4:  # Top-right corner  
                cell.border = top_right_border
            elif row_num == last_row and col_num == 1:  # Bottom-left corner
                cell.border = bottom_left_border
            elif row_num == last_row and col_num == 4:  # Bottom-right corner
                cell.border = bottom_right_border
            elif row_num == 1:  # Top edge
                cell.border = top_border
            elif row_num == last_row:  # Bottom edge
                cell.border = bottom_border
            elif col_num == 1:  # Left edge
                cell.border = left_border
            elif col_num == 4:  # Right edge
                cell.border = right_border
            else:  # Internal cells
                cell.border = normal_border
    
    # Aggiungi padding alle righe (altezza minima)
    for row_num in range(1, last_row + 1):
        ws.row_dimensions[row_num].height = 25  # Altezza maggiore per più spazio
    
    # Autofit colonne con padding aggiuntivo per spaziatura
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Aumenta la larghezza per più spaziatura
        adjusted_width = min(max_length + 4, 50)  # +4 invece di +2 per più spazio
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=seriali_ordine_{order_number}.xlsx"}
    )

@router.get("/export-all-excel")
async def export_all_serials_excel(db: Session = Depends(get_db)):
    """
    Esporta tutti i seriali del sistema in formato Excel usando il template
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export non disponibile. Installare openpyxl.")
    
    # Carica il template Excel
    try:
        from openpyxl import load_workbook
        template_path = "/mnt/c/WMS_EPM/Esempio file excel/Template Seriali.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Rinomina il sheet
        ws.title = "Tutti i Seriali"
        
    except Exception as e:
        # Se il template non è disponibile, crea un nuovo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tutti i Seriali"
        
        # Stili del template
        from openpyxl.styles import Font, Alignment
        header_font = Font(name="Aptos Narrow", bold=True, size=11)
        header_alignment = Alignment(horizontal='center')
        
        # Crea header manualmente
        headers = ['Ordine', 'Codice Prodotto', 'Seriale prodotto', 'Data']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
    
    # Query per ottenere tutti i seriali
    serials_query = db.query(models.serials.ProductSerial).order_by(
        models.serials.ProductSerial.order_number,
        models.serials.ProductSerial.product_sku,
        models.serials.ProductSerial.serial_number
    ).all()
    
    if not serials_query:
        raise HTTPException(status_code=404, detail="Nessun seriale trovato nel sistema")
    
    # Popola i dati partendo dalla riga 2
    row = 2
    for serial in serials_query:
        ws.cell(row=row, column=1, value=serial.order_number)  # Ordine
        ws.cell(row=row, column=2, value=serial.product_sku)   # Codice Prodotto
        ws.cell(row=row, column=3, value=serial.serial_number) # Seriale prodotto
        
        # Data formattata - solo data senza orario
        if serial.uploaded_at:
            date_formatted = serial.uploaded_at.strftime("%d/%m/%Y")
            ws.cell(row=row, column=4, value=date_formatted)
        elif serial.created_at:
            date_formatted = serial.created_at.strftime("%d/%m/%Y")
            ws.cell(row=row, column=4, value=date_formatted)
        else:
            ws.cell(row=row, column=4, value="N/A")
        
        row += 1
    
    # Autofit colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Nome file con data
    current_date = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"tutti_seriali_{current_date}.xlsx"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/format-info", response_model=schemas.serials.SerialFileFormat)
def get_file_format_info():
    """Informazioni sul formato file seriali"""
    return schemas.serials.SerialFileFormat()

@router.get("/duplicates")
def get_duplicate_serials(db: Session = Depends(get_db)):
    """Trova tutti i seriali duplicati nel sistema"""
    serial_service = SerialService(db)
    return serial_service.get_duplicate_serials_in_system()

@router.get("/check/{serial_number}")
def check_serial_exists(serial_number: str, db: Session = Depends(get_db)):
    """Verifica se un seriale specifico esiste già"""
    serial_service = SerialService(db)
    existing = serial_service.check_serial_exists(serial_number)
    
    if existing:
        return {
            "exists": True,
            "serial_number": serial_number,
            "order_number": existing.order_number,
            "product_sku": existing.product_sku,
            "ean_code": existing.ean_code,
            "uploaded_at": existing.uploaded_at
        }
    else:
        return {
            "exists": False,
            "serial_number": serial_number
        }

@router.delete("/orders/{order_number}")
def delete_order_serials(order_number: str, db: Session = Depends(get_db)):
    """Elimina tutti i seriali di un ordine"""
    # Log dell'operazione prima della cancellazione
    logger = LoggingService(db)
    
    deleted_count = db.query(models.serials.ProductSerial).filter(
        models.serials.ProductSerial.order_number == order_number
    ).delete()
    
    # Log operazione di cancellazione
    if deleted_count > 0:
        logger.log_operation(
            operation_type=OperationType.SERIALI_RIMOSSI,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=None,  # Vuoto come richiesto
            quantity=deleted_count,
            details={
                "order_number": order_number,
                "deleted_count": deleted_count,
                "operation": "delete_order_serials"
            }
        )
    
    db.commit()
    
    return {"message": f"Eliminati {deleted_count} seriali per ordine {order_number}"}

@router.delete("/batch/{upload_batch_id}")
def delete_batch_serials(upload_batch_id: str, db: Session = Depends(get_db)):
    """Elimina tutti i seriali di un batch upload"""
    # Log dell'operazione prima della cancellazione
    logger = LoggingService(db)
    
    deleted_count = db.query(models.serials.ProductSerial).filter(
        models.serials.ProductSerial.upload_batch_id == upload_batch_id
    ).delete()
    
    # Log operazione di cancellazione
    if deleted_count > 0:
        logger.log_operation(
            operation_type=OperationType.SERIALI_RIMOSSI,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=None,  # Vuoto come richiesto
            quantity=deleted_count,
            details={
                "upload_batch_id": upload_batch_id,
                "deleted_count": deleted_count,
                "operation": "delete_batch_serials"
            }
        )
    
    db.commit()
    
    return {"message": f"Eliminati {deleted_count} seriali per batch {upload_batch_id}"}

@router.post("/parse-file", response_model=schemas.serials.SerialParseResult)
async def parse_serial_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Parsing del file seriali con recap modificabile
    Restituisce anteprima senza eseguire commit
    """
    # Verifica formato file
    if not file.filename.endswith(('.txt', '.csv')):
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usare .txt o .csv")
    
    try:
        # Leggi contenuto file
        content = await file.read()
        file_content = content.decode('utf-8')
        
        # Processa con SerialService (solo parsing, no commit)
        serial_service = SerialService(db)
        result = serial_service.parse_serial_file_with_recap(file_content, file.filename)
        
        return result
        
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Errore codifica file. Usare codifica UTF-8.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante il parsing: {str(e)}")

@router.post("/commit-operations", response_model=schemas.serials.SerialUploadResult)
async def commit_serial_operations(
    request: schemas.serials.SerialCommitRequest,
    db: Session = Depends(get_db)
):
    """
    Commit delle operazioni seriali dopo validazione recap
    """
    try:
        serial_service = SerialService(db)
        result = serial_service.commit_serial_operations(request)
        
        if not result.success:
            raise HTTPException(status_code=400, detail=result.message)
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante il commit: {str(e)}")