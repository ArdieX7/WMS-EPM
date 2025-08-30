from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.sql import func
from typing import List, Dict, Tuple, Any
from collections import defaultdict
from datetime import datetime
import os
import shutil
from pathlib import Path

from wms_app import models, schemas
from wms_app.database import database, get_db
from wms_app.routers.auth import require_permission
from wms_app.services.logging_service import LoggingService
from wms_app.models.logs import OperationType, OperationCategory, OperationStatus

# Import templates in modo lazy per evitare import circolari
def get_templates():
    from wms_app.main import templates
    return templates

router = APIRouter(
    prefix="/orders",
    tags=["orders"],
)

# --- Viste HTML (devono essere definite prima delle rotte con parametri di percorso) ---

@router.get("/manage", response_class=HTMLResponse)
async def get_orders_management_page(request: Request, db: Session = Depends(get_db)):
    # Escludi ordini archiviati dalla vista principale
    orders = db.query(models.Order).filter(models.Order.is_archived == 0).options(joinedload(models.Order.lines)).all()
    products = db.query(models.Product).all()
    return get_templates().TemplateResponse("orders.html", {
        "request": request,
        "orders": orders,
        "products": products,
        "active_page": "orders"
    })

# --- API Endpoints per la Gestione Ordini (generici, senza parametri di percorso) ---

@router.post("/", response_model=schemas.Order)
def create_order(order: schemas.OrderCreate, db: Session = Depends(get_db)):
    logger = LoggingService(db)
    
    # Controlla ordine duplicato
    db_order = db.query(models.Order).filter(models.Order.order_number == order.order_number).first()
    if db_order:
        # Log errore ordine duplicato
        logger.log_error(
            operation_type=OperationType.ORDINE_CREATO,
            error=f"Order number {order.order_number} already exists",
            operation_category=OperationCategory.MANUAL,
            file_name=f"ORDER_{order.order_number}",
            details={
                'order_number': order.order_number,
                'customer_name': order.customer_name,
                'error_reason': 'duplicate_order_number',
                'operation_description': f"Tentativo fallito creazione ordine {order.order_number}: numero ordine già esistente"
            },
            api_endpoint="/orders/"
        )
        raise HTTPException(status_code=400, detail="Order number already exists")

    new_order = models.Order(order_number=order.order_number, customer_name=order.customer_name)
    db.add(new_order)
    
    try:
        db.flush() # Per ottenere l'ID dell'ordine prima del commit
    except Exception as e:
        logger.log_error(
            operation_type=OperationType.ORDINE_CREATO,
            error=e,
            operation_category=OperationCategory.MANUAL,
            file_name=f"ORDER_{order.order_number}",
            details={
                'order_number': order.order_number,
                'customer_name': order.customer_name,
                'error_reason': 'database_flush_failed',
                'operation_description': f"Errore database durante creazione ordine {order.order_number}"
            },
            api_endpoint="/orders/"
        )
        raise HTTPException(status_code=500, detail="Database error during order creation")

    for line_data in order.lines:
        product = db.query(models.Product).filter(models.Product.sku == line_data.product_sku).first()
        if not product:
            # Log errore prodotto non trovato
            logger.log_error(
                operation_type=OperationType.ORDINE_CREATO,
                error=f"Product SKU {line_data.product_sku} not found",
                operation_category=OperationCategory.MANUAL,
                product_sku=line_data.product_sku,
                file_name=f"ORDER_{order.order_number}",
                details={
                    'order_number': order.order_number,
                    'customer_name': order.customer_name,
                    'error_reason': 'product_not_found',
                    'missing_sku': line_data.product_sku,
                    'operation_description': f"Errore creazione ordine {order.order_number}: prodotto {line_data.product_sku} non trovato"
                },
                api_endpoint="/orders/"
            )
            raise HTTPException(status_code=404, detail=f"Product SKU {line_data.product_sku} not found for order line")
        
        new_line = models.OrderLine(
            order_id=new_order.id,
            product_sku=line_data.product_sku,
            requested_quantity=line_data.requested_quantity
        )
        db.add(new_line)
    
    # LOGGING: Registra la creazione manuale dell'ordine (successo)
    
    # Logga ogni prodotto dell'ordine separatamente per visibilità nelle colonne
    for line_data in order.lines:
        logger.log_operation(
            operation_type=OperationType.ORDINE_CREATO,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=line_data.product_sku,  # SKU del prodotto nell'ordine
            quantity=line_data.requested_quantity,  # Quantità richiesta
            user_id="manual_user",  # TODO: Sostituire con sistema auth reale
            file_name=f"ORDER_{new_order.order_number}",  # Numero ordine nella colonna Dettagli
            details={
                'order_number': new_order.order_number,  # Numero ordine nei dettagli JSON
                'customer_name': new_order.customer_name,
                'creation_method': 'manual',
                'operation_description': f"Creazione manuale ordine {new_order.order_number}: aggiunto {line_data.requested_quantity}x {line_data.product_sku} per cliente {new_order.customer_name}",
                'total_order_lines': len(order.lines),
                'total_order_items': sum(line.requested_quantity for line in order.lines)
            },
            api_endpoint="/orders/"
        )
    
    db.commit()
    db.refresh(new_order)
    return new_order

@router.get("/", response_model=List[schemas.Order])
def read_orders(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    # Escludi ordini archiviati dalla lista principale
    orders = db.query(models.Order).filter(
        models.Order.is_archived == 0
    ).options(
        joinedload(models.Order.lines).joinedload(models.OrderLine.product)
    ).offset(skip).limit(limit).all()
    
    # Calcola il peso totale per ogni ordine
    for order in orders:
        total_weight = 0.0
        for line in order.lines:
            if line.product and line.product.weight:
                total_weight += line.requested_quantity * line.product.weight
        order.total_weight = total_weight
    
    return orders

@router.get("/archived")
def get_archived_orders(db: Session = Depends(get_db)):
    """Recupera tutti gli ordini archiviati."""
    try:
        archived_orders = db.query(models.Order).filter(
            models.Order.is_archived == 1
        ).options(joinedload(models.Order.lines).joinedload(models.OrderLine.product)).order_by(models.Order.id.desc()).all()
        
        orders_data = []
        for order in archived_orders:
            # Calcola il peso totale per ogni ordine
            total_weight = 0.0
            for line in order.lines:
                if line.product and line.product.weight:
                    total_weight += line.requested_quantity * line.product.weight
            
            order_data = {
                "id": order.id,
                "order_number": order.order_number,
                "customer_name": order.customer_name,
                "order_date": order.order_date.isoformat() if order.order_date else None,
                "archived_date": order.archived_date.isoformat() if order.archived_date else None,
                "is_completed": bool(order.is_completed),
                "is_cancelled": bool(order.is_cancelled),
                "ddt_number": order.ddt_number,
                "total_weight": total_weight,
                "lines": [
                    {
                        "product_sku": line.product_sku,
                        "requested_quantity": line.requested_quantity,
                        "picked_quantity": line.picked_quantity
                    }
                    for line in order.lines
                ]
            }
            orders_data.append(order_data)
        
        return {"orders": orders_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching archived orders: {str(e)}")

# --- API Endpoints con parametri di percorso (devono essere definiti dopo le rotte generiche) ---


@router.post("/import-orders-txt")
async def import_orders_from_txt(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    try:
        text_content = content.decode("utf-8")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Il file non è in formato UTF-8 valido.")

    orders_in_file = {}
    line_number = 0
    for line in text_content.splitlines():
        line_number += 1
        if not line.strip():
            continue
        
        parts = line.strip().split(',')
        if len(parts) != 4:
            raise HTTPException(status_code=400, detail=f"Formato non valido alla riga {line_number}: la riga deve contenere NumeroOrdine,Cliente,SKU,Qty")

        order_number, customer_name, sku, qty_str = parts
        order_number = order_number.strip()
        customer_name = customer_name.strip()
        sku = sku.strip()

        try:
            quantity = int(qty_str.strip())
            if quantity <= 0:
                raise ValueError("La quantità deve essere positiva.")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Quantità non valida alla riga {line_number}: '{qty_str}'")

        if order_number not in orders_in_file:
            orders_in_file[order_number] = {
                "customer_name": customer_name,
                "lines": []
            }
        
        orders_in_file[order_number]["lines"].append({"sku": sku, "quantity": quantity})

    created_count = 0
    updated_count = 0

    for order_number, order_data in orders_in_file.items():
        db_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()

        if not db_order:
            # Crea un nuovo ordine
            db_order = models.Order(
                order_number=order_number, 
                customer_name=order_data['customer_name']
            )
            db.add(db_order)
            db.flush() # Per ottenere l'ID dell'ordine
            created_count += 1
        else:
            # Ordine esistente, verifica se è già completato
            if db_order.is_completed:
                continue # Salta gli ordini già completati
            updated_count += 1

        # Aggiungi le righe d'ordine
        for line in order_data['lines']:
            # Controlla se il prodotto esiste
            product = db.query(models.Product).filter(models.Product.sku == line['sku']).first()
            if not product:
                db.rollback()
                raise HTTPException(status_code=404, detail=f"Prodotto con SKU '{line['sku']}' non trovato per l'ordine '{order_number}'. L'importazione è stata annullata.")

            # Controlla se una riga simile esiste già per evitare duplicati
            existing_line = db.query(models.OrderLine).filter(
                models.OrderLine.order_id == db_order.id,
                models.OrderLine.product_sku == line['sku']
            ).first()

            if existing_line:
                # Aggiorna la quantità della riga esistente
                existing_line.requested_quantity += line['quantity']
            else:
                # Crea una nuova riga d'ordine
                new_line = models.OrderLine(
                    order_id=db_order.id,
                    product_sku=line['sku'],
                    requested_quantity=line['quantity']
                )
                db.add(new_line)

    # LOGGING: Registra l'import da file TXT
    logger = LoggingService(db)
    file_name = file.filename if hasattr(file, 'filename') else 'orders_import.txt'
    
    # Logga ogni prodotto di ogni ordine importato
    for order_number, order_data in orders_in_file.items():
        db_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
        if db_order:  # Solo se l'ordine è stato creato/aggiornato con successo
            for line in order_data['lines']:
                operation_type_to_use = OperationType.ORDINE_CREATO  # Default per ordini nuovi
                creation_method = 'file_import'
                
                # Determina se è creazione o aggiornamento
                if order_number in [o for o in orders_in_file.keys()]:
                    existing_order_check = db.query(models.Order).filter(
                        models.Order.order_number == order_number,
                        models.Order.id != db_order.id
                    ).first()
                    if existing_order_check:
                        operation_type_to_use = OperationType.ORDINE_MODIFICATO
                        creation_method = 'file_update'
                
                logger.log_operation(
                    operation_type=operation_type_to_use,
                    operation_category=OperationCategory.FILE,
                    status=OperationStatus.SUCCESS,
                    product_sku=line['sku'],  # SKU del prodotto importato
                    quantity=line['quantity'],  # Quantità richiesta
                    user_id="file_import_user",
                    file_name=f"ORDER_{order_number}",  # Numero ordine nella colonna Dettagli
                    details={
                        'order_number': order_number,  # Numero ordine nei dettagli JSON
                        'customer_name': order_data['customer_name'],
                        'creation_method': creation_method,
                        'source_file': file_name,
                        'operation_description': f"Import file: {creation_method} ordine {order_number}, aggiunto {line['quantity']}x {line['sku']} per cliente {order_data['customer_name']}",
                        'import_stats': {
                            'orders_created': created_count,
                            'orders_updated': updated_count
                        }
                    },
                    api_endpoint="/orders/import-orders-txt"
                )

    db.commit()
    return {"message": f"Importazione completata. Ordini creati: {created_count}. Ordini aggiornati: {updated_count}."}

# --- Nuovi Endpoint Picking (DEVONO essere prima di /{order_id}) ---

@router.post("/validate-picking-txt")
async def validate_picking_from_txt(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Valida le operazioni di picking da file TXT scanner e restituisce un recap dettagliato.
    Formato pistola scanner: numero ordine, ubicazione, EAN/SKU ripetuti o EAN/SKU_quantità
    Compatibile con il sistema recap dell'inventario per editing pre-commit.
    """
    try:
        picking_data, parse_errors = await _parse_picking_file_scanner(file, db)
    except HTTPException as e:
        raise e
    
    recap_items = []
    errors = []
    warnings = []
    line_counter = 0
    
    # Prepara mappe per validazione
    all_orders = {o.order_number: o for o in db.query(models.Order).options(joinedload(models.Order.lines)).all()}
    all_products = {p.sku: p for p in db.query(models.Product).all()}
    all_eans = {e.ean: e.product_sku for e in db.query(models.EanCode).all()}
    
    # Aggiungi errori di parsing al recap
    for parse_error in parse_errors:
        errors.append({
            "line": parse_error["line"],
            "message": parse_error["message"],
            "field": parse_error["field"],
            "value": parse_error["value"]
        })
        
        # Crea un recap item per permettere la correzione
        line_counter += 1
        
        # Determina customer_name se l'ordine esiste
        customer_name = ""
        if parse_error["order"] != "MANCANTE":
            order = all_orders.get(parse_error["order"])
            if order:
                customer_name = order.customer_name
        
        # Logica migliorata per i campi
        location = ""
        sku = ""
        
        if parse_error["field"] == "location":
            # Manca solo l'ubicazione, il prodotto è specificato
            location = ""  # Campo vuoto da compilare
            sku = parse_error["value"]  # Mantieni il SKU dal file
        elif parse_error["field"] == "sku":
            # Prodotto non trovato nel database, ubicazione è ok
            location = parse_error["location"]
            sku = parse_error["value"]  # Mantieni quello dal file per editing
        elif parse_error["field"] == "order":
            # Ordine mancante
            location = parse_error["location"] if parse_error["location"] != "MANCANTE" else ""
            sku = parse_error["value"]
        else:
            # Caso generale
            location = parse_error["location"] if parse_error["location"] != "MANCANTE" else ""
            sku = parse_error["value"]
        
        recap_item = {
            "line": parse_error["line"],
            "order_number": parse_error["order"] if parse_error["order"] != "MANCANTE" else "",
            "customer_name": customer_name,
            "location": location,
            "sku": sku,
            "description": "",
            "input_code": parse_error["value"],
            "quantity": 1,
            "current_stock": 0,
            "remaining_to_pick": 0,
            "remaining_stock": 0,
            "status": "error"
        }
        recap_items.append(recap_item)
    
    # Processa tutti i dati parsati correttamente
    for order_number, locations_data in picking_data.items():
        order = all_orders.get(order_number)
        
        for location, skus_data in locations_data.items():
            for sku, quantity in skus_data.items():
                line_counter += 1
                
                # Prepara item recap
                recap_item = {
                    "line": line_counter,
                    "order_number": order_number,
                    "customer_name": order.customer_name if order else "ORDINE NON TROVATO",
                    "location": location,
                    "sku": sku,
                    "description": "",
                    "input_code": sku,  # Codice originale dal file
                    "quantity": quantity,
                    "current_stock": 0,
                    "remaining_to_pick": 0,
                    "remaining_stock": 0,
                    "status": "ok"
                }
                
                # Validazione ordine
                if not order:
                    errors.append({
                        "line": line_counter,
                        "message": f"Ordine '{order_number}' non trovato nel database",
                        "field": "order_number",
                        "value": order_number
                    })
                    recap_item["status"] = "error"
                    recap_item["customer_name"] = "ORDINE NON TROVATO"
                    
                elif order.is_completed:
                    errors.append({
                        "line": line_counter,
                        "message": f"Ordine '{order_number}' è già completato",
                        "field": "order",
                        "value": order_number
                    })
                    recap_item["status"] = "error"
                    
                else:
                    # Validazione prodotto nell'ordine
                    order_line = None
                    for line in order.lines:
                        if line.product_sku == sku:
                            order_line = line
                            break
                    
                    if not order_line:
                        # Controlla se è un EAN code che corrisponde a un SKU nell'ordine
                        actual_sku = all_eans.get(sku)
                        if actual_sku:
                            for line in order.lines:
                                if line.product_sku == actual_sku:
                                    order_line = line
                                    recap_item["sku"] = actual_sku  # Aggiorna con il SKU corretto
                                    break
                    
                    if not order_line:
                        errors.append({
                            "line": line_counter,
                            "message": f"Prodotto '{sku}' non presente nell'ordine {order_number}",
                            "field": "sku",
                            "value": sku
                        })
                        recap_item["status"] = "error"
                    else:
                        # Calcola quantità rimanente da prelevare dall'ordine
                        remaining_order = order_line.requested_quantity - order_line.picked_quantity
                        recap_item["remaining_to_pick"] = remaining_order
                        
                        if quantity > remaining_order:
                            if remaining_order == 0:
                                warnings.append({
                                    "line": line_counter,
                                    "message": f"Prodotto '{sku}' già completamente prelevato per ordine {order_number}",
                                    "field": "quantity",
                                    "value": quantity
                                })
                                recap_item["status"] = "warning"
                            else:
                                warnings.append({
                                    "line": line_counter,
                                    "message": f"Quantità eccessiva: rimanenti da prelevare {remaining_order}, tentativo prelievo {quantity}",
                                    "field": "quantity", 
                                    "value": quantity
                                })
                                recap_item["status"] = "warning"
                
                # Validazione giacenza
                inventory_item = db.query(models.Inventory).filter(
                    models.Inventory.location_name == location,
                    models.Inventory.product_sku == recap_item["sku"]
                ).first()
                
                if inventory_item:
                    recap_item["current_stock"] = inventory_item.quantity
                    # Calcola giacenza rimanente dopo il prelievo
                    recap_item["remaining_stock"] = max(0, inventory_item.quantity - quantity)
                    if inventory_item.quantity < quantity:
                        errors.append({
                            "line": line_counter,
                            "message": f"Giacenza insufficiente in '{location}': disponibili {inventory_item.quantity}, richiesti {quantity}",
                            "field": "quantity",
                            "value": quantity
                        })
                        recap_item["status"] = "error"
                else:
                    recap_item["current_stock"] = 0
                    recap_item["remaining_stock"] = 0
                    errors.append({
                        "line": line_counter,
                        "message": f"Prodotto '{recap_item['sku']}' non presente in ubicazione '{location}'",
                        "field": "location",
                        "value": location
                    })
                    recap_item["status"] = "error"
                
                # Aggiungi descrizione prodotto se disponibile
                product = all_products.get(recap_item["sku"])
                if product:
                    recap_item["description"] = product.description
                
                recap_items.append(recap_item)
    
    # Prepara risultato finale in formato compatibile con recap inventory
    result = {
        "recap_items": recap_items,
        "errors": errors,
        "warnings": warnings,
        "stats": {
            "total": len(recap_items),
            "ok": len([item for item in recap_items if item["status"] == "ok"]),
            "warnings": len([item for item in recap_items if item["status"] == "warning"]),
            "errors": len([item for item in recap_items if item["status"] == "error"])
        },
        "orders_summary": {}
    }
    
    # Aggiungi summary per ordine
    for order_number in picking_data.keys():
        order = all_orders.get(order_number)
        if order:
            order_items = [item for item in recap_items if item["order_number"] == order_number]
            result["orders_summary"][order_number] = {
                "customer_name": order.customer_name,
                "is_completed": order.is_completed,
                "total_operations": len(order_items),
                "valid_operations": len([item for item in order_items if item["status"] == "ok"]),
                "lines": {line.product_sku: {
                    "requested": line.requested_quantity,
                    "picked": line.picked_quantity,
                    "remaining": line.requested_quantity - line.picked_quantity
                } for line in order.lines}
            }
    
    return result

@router.post("/commit-picking-txt")
async def commit_picking_from_txt(file: UploadFile = File(...), force: bool = False, db: Session = Depends(get_db)):
    """
    Commit delle operazioni di picking dopo validazione.
    Se force=True, esegue solo le operazioni valide ignorando quelle con warning/errori.
    """
    try:
        picking_data, parse_errors = await _parse_picking_file_scanner(file, db)
    except HTTPException as e:
        raise e
    
    if not picking_data and not parse_errors:
        raise HTTPException(status_code=400, detail="Il file è vuoto o non contiene dati validi.")
    
    if parse_errors and not force:
        raise HTTPException(status_code=400, detail=f"Errori di parsing trovati. Usa force=true per ignorarli: {len(parse_errors)} errori")
    
    successful_operations = []
    skipped_operations = []
    
    for order_number, locations_data in picking_data.items():
        # Trova l'ordine
        order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
        if not order or order.is_completed:
            skipped_operations.append(f"Ordine '{order_number}' saltato (non trovato o completato)")
            continue
        
        for location, skus_data in locations_data.items():
            for sku, quantity in skus_data.items():
                # Trova la riga ordine
                order_line = db.query(models.OrderLine).filter(
                    models.OrderLine.order_id == order.id,
                    models.OrderLine.product_sku == sku
                ).first()
                
                if not order_line:
                    skipped_operations.append(f"Prodotto '{sku}' non trovato nell'ordine '{order_number}'")
                    continue
                
                # Verifica giacenza disponibile
                inventory_item = db.query(models.Inventory).filter(
                    models.Inventory.location_name == location,
                    models.Inventory.product_sku == sku
                ).first()
                
                if not inventory_item or inventory_item.quantity < quantity:
                    skipped_operations.append(f"Giacenza insufficiente per {sku} in {location}")
                    continue
                
                # Verifica quantità ordine (con tolleranza se force=True)
                remaining_to_pick = order_line.requested_quantity - order_line.picked_quantity
                actual_quantity = quantity
                
                if quantity > remaining_to_pick:
                    if force and remaining_to_pick > 0:
                        actual_quantity = remaining_to_pick  # Preleva solo quello che serve
                        skipped_operations.append(f"Ridotta quantità per {sku} da {quantity} a {actual_quantity}")
                    elif remaining_to_pick == 0:
                        skipped_operations.append(f"Saltato {sku}: già completamente prelevato")
                        continue
                    else:
                        skipped_operations.append(f"Saltato {sku}: quantità eccessiva")
                        continue
                
                # Esegui l'operazione di picking
                inventory_item.quantity -= actual_quantity
                order_line.picked_quantity += actual_quantity
                
                # Aggiungi a OutgoingStock
                outgoing_item = db.query(models.OutgoingStock).filter(
                    models.OutgoingStock.order_line_id == order_line.id,
                    models.OutgoingStock.product_sku == sku
                ).first()
                
                if outgoing_item:
                    outgoing_item.quantity += actual_quantity
                else:
                    new_outgoing_item = models.OutgoingStock(
                        order_line_id=order_line.id,
                        product_sku=sku,
                        quantity=actual_quantity
                    )
                    db.add(new_outgoing_item)
                
                successful_operations.append(f"Prelevato {actual_quantity}x {sku} da {location} per ordine {order_number}")
    
    if not successful_operations and not force:
        db.rollback()
        raise HTTPException(status_code=400, detail="Nessuna operazione valida da eseguire")
    
    # LOGGING: Registra le operazioni di picking da file
    logger = LoggingService(db)
    
    # Prepara operazioni per logging
    batch_operations = []
    for order_number, locations_data in picking_data.items():
        for location, skus_data in locations_data.items():
            for sku, quantity in skus_data.items():
                # Controlla se l'operazione è stata eseguita con successo
                operation_found = any(f"{quantity}x {sku} da {location} per ordine {order_number}" in op 
                                    for op in successful_operations)
                if operation_found:
                    batch_operations.append({
                        'product_sku': sku,
                        'location_from': location,
                        'location_to': None,  # Picking: scala da inventario
                        'quantity': quantity,
                        'status': OperationStatus.SUCCESS,
                        'details': {
                            'order_number': order_number,
                            'operation_description': f"Picking da file: {sku} ({quantity} pz) da {location} per ordine {order_number}",
                            'source': 'picking_file_scanner',
                            'picking_type': 'file_picking'
                        }
                    })
    
    # Registra operazioni senza log batch start/end
    if batch_operations:
        file_name = file.filename if hasattr(file, 'filename') else 'picking_file.txt'
        logger.log_file_operations(
            operation_type=OperationType.PRELIEVO_FILE,
            operation_category=OperationCategory.FILE,
            operations=batch_operations,
            file_name=file_name,
            user_id="file_user"
        )
    
    db.commit()
    
    return {
        "message": f"Picking completato: {len(successful_operations)} operazioni eseguite",
        "successful_operations": successful_operations,
        "skipped_operations": skipped_operations,
        "force_mode": force
    }

@router.post("/debug-picking-txt")
async def debug_picking_from_txt(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Endpoint di debug per verificare il parsing del file scanner.
    """
    try:
        picking_data, parse_errors = await _parse_picking_file_scanner(file, db)
        
        # Check database contents
        all_locations = {loc.name for loc in db.query(models.Location).all()}
        all_products = {p.sku for p in db.query(models.Product).all()}
        all_eans = {e.ean: e.product_sku for e in db.query(models.EanCode).all()}
        
        return {
            "file_name": file.filename,
            "parsed_data": dict(picking_data),
            "parse_errors": parse_errors,
            "parser_used": "new_scanner_parser",
            "database_info": {
                "locations": list(all_locations)[:10],
                "products": list(all_products)[:10], 
                "eans": list(all_eans.keys())[:10]
            }
        }
    except Exception as e:
        return {"error": str(e), "type": type(e).__name__}

@router.post("/import-picking-txt")
async def import_picking_from_txt_legacy(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Endpoint legacy che usa il nuovo parser scanner ma mantiene il comportamento originale.
    """
    try:
        # Parsing con il nuovo formato scanner
        picking_data, parse_errors = await _parse_picking_file_scanner(file, db)
        
        if not picking_data and not parse_errors:
            raise HTTPException(status_code=400, detail="Il file è vuoto o non contiene dati validi.")
        
        if parse_errors:
            # Endpoint legacy fallisce con errori di parsing
            error_messages = [f"Riga {e['line']}: {e['message']}" for e in parse_errors]
            raise HTTPException(status_code=400, detail=f"Errori nel parsing: {'; '.join(error_messages)}")
        
        return {"message": f"Parser chiamato correttamente. Ordini trovati: {list(picking_data.keys())}"}
        
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore interno: {str(e)}")

# --- API Endpoints con parametri di percorso (devono essere definiti dopo le rotte generiche) ---

@router.get("/{order_id}", response_model=schemas.Order)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).options(joinedload(models.Order.lines)).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order

# --- Logica di Picking ---

@router.get("/{order_id}/picking-suggestions", response_model=Dict[str, schemas.PickingSuggestion])
def get_picking_suggestions(order_id: int, db: Session = Depends(get_db)):
    from wms_app.services.reservation_service import ReservationService
    
    order = db.query(models.Order).options(joinedload(models.Order.lines)).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.is_completed:
        raise HTTPException(status_code=400, detail="Order is already completed")

    # Inizializza il servizio di prenotazioni
    reservation_service = ReservationService(db)
    
    # Pulisce automaticamente le prenotazioni scadute
    reservation_service.cleanup_expired_reservations()
    
    # Prepara lista prodotti necessari per l'ordine
    products_needed = []
    for line in order.lines:
        if line.requested_quantity > line.picked_quantity:
            remaining_to_pick = line.requested_quantity - line.picked_quantity
            products_needed.append({
                'sku': line.product_sku,
                'quantity': remaining_to_pick,
                'line_id': line.id
            })
    
    if not products_needed:
        return {}  # Ordine già completamente prelevato
    
    # Usa il Round-Robin Reservation System per allocare ubicazioni
    try:
        allocations = reservation_service.allocate_picking_locations(
            order_id=str(order.order_number), 
            products_needed=products_needed
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore allocazione picking: {str(e)}")
    
    # Converte il risultato nel formato atteso dal frontend
    suggestions = {}
    for allocation in allocations:
        sku = allocation['sku']
        
        product_suggestions = []
        for loc_allocation in allocation['allocations']:
            product_suggestions.append({
                "location_name": loc_allocation['location_name'],
                "quantity": loc_allocation['quantity'],
                "reservation_id": loc_allocation['reservation_id']  # Nuovo campo per tracking
            })
        
        if allocation['fully_allocated']:
            suggestions[sku] = schemas.PickingSuggestion(
                status="full_stock",
                needed=allocation['requested_quantity'],
                available_in_locations=product_suggestions
            )
        elif len(allocation['allocations']) > 0:
            # Stock parziale disponibile
            suggestions[sku] = schemas.PickingSuggestion(
                status="partial_stock",
                needed=allocation['requested_quantity'],
                available_in_locations=product_suggestions
            )
        else:
            # Nessuna allocazione disponibile - prodotto non presente in inventario
            suggestions[sku] = schemas.PickingSuggestion(
                status="out_of_stock",
                needed=allocation['requested_quantity'],
                available_in_locations=[]
            )
    
    return suggestions

@router.post("/{order_id}/confirm-pick", response_model=schemas.Order)
def confirm_pick(order_id: int, pick_confirmation: schemas.PickConfirmation, db: Session = Depends(get_db)):
    from wms_app.services.reservation_service import ReservationService
    
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.is_completed:
        raise HTTPException(status_code=400, detail="Order is already completed")

    reservation_service = ReservationService(db)

    for picked_item in pick_confirmation.picked_items:
        order_line = db.query(models.OrderLine).filter(
            models.OrderLine.id == picked_item.order_line_id,
            models.OrderLine.order_id == order_id
        ).first()
        if not order_line:
            raise HTTPException(status_code=404, detail=f"Order line {picked_item.order_line_id} not found for this order")

        inventory_item = db.query(models.Inventory).filter(
            models.Inventory.product_sku == picked_item.product_sku,
            models.Inventory.location_name == picked_item.location_name
        ).first()

        if not inventory_item or inventory_item.quantity < picked_item.quantity:
            raise HTTPException(status_code=400, detail=f"Not enough stock of {picked_item.product_sku} in {picked_item.location_name} to pick {picked_item.quantity}")

        # Scala dalla giacenza
        inventory_item.quantity -= picked_item.quantity
        order_line.picked_quantity += picked_item.quantity

        # Sposta in OutgoingStock
        outgoing_item = db.query(models.OutgoingStock).filter(
            models.OutgoingStock.order_line_id == picked_item.order_line_id,
            models.OutgoingStock.product_sku == picked_item.product_sku
        ).first()

        if outgoing_item:
            outgoing_item.quantity += picked_item.quantity
        else:
            new_outgoing_item = models.OutgoingStock(
                order_line_id=picked_item.order_line_id,
                product_sku=picked_item.product_sku,
                quantity=picked_item.quantity
            )
            db.add(new_outgoing_item)
        
        # NUOVO: Completa la prenotazione se present
        # Cerca prenotazioni attive per questo ordine/prodotto/ubicazione
        if hasattr(picked_item, 'reservation_id') and picked_item.reservation_id:
            reservation_service.complete_reservation(picked_item.reservation_id, picked_item.quantity)
        else:
            # Fallback: cerca prenotazione per order_number/sku/location
            from wms_app.models.reservations import InventoryReservation
            reservation = db.query(InventoryReservation).filter(
                InventoryReservation.order_id == str(order.order_number),
                InventoryReservation.product_sku == picked_item.product_sku,
                InventoryReservation.location_name == picked_item.location_name,
                InventoryReservation.status == 'active'
            ).first()
            
            if reservation:
                reservation_service.complete_reservation(reservation.id, picked_item.quantity)
    
    # LOGGING: Registra le operazioni di picking manuale
    logger = LoggingService(db)
    
    # Crea operazioni di log per ogni item prelevato
    for picked_item in pick_confirmation.picked_items:
        logger.log_operation(
            operation_type=OperationType.PRELIEVO_MANUALE,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=picked_item.product_sku,
            location_from=picked_item.location_name,
            location_to=None,  # Picking: scala da inventario
            quantity=picked_item.quantity,
            user_id="picking_user",  # TODO: Sostituire con sistema auth reale
            details={
                'order_number': order.order_number,
                'order_line_id': picked_item.order_line_id,
                'operation_description': f"Picking manuale: {picked_item.product_sku} ({picked_item.quantity} pz) da {picked_item.location_name} per ordine {order.order_number}",
                'picking_type': 'manual_picking',
                'reservation_id': getattr(picked_item, 'reservation_id', None),
                'customer_name': order.customer_name
            },
            api_endpoint="/orders/{order_id}/confirm-pick"
        )
    
    db.commit()
    db.refresh(order)
    return order

@router.post("/{order_id}/fulfill", response_model=schemas.Order)
def fulfill_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.is_completed:
        raise HTTPException(status_code=400, detail="Order is already completed")

    # Scala da OutgoingStock e completa l'ordine
    for line in order.lines:
        outgoing_item = db.query(models.OutgoingStock).filter(
            models.OutgoingStock.order_line_id == line.id,
            models.OutgoingStock.product_sku == line.product_sku
        ).first()

        if outgoing_item:
            # Sincronizza OutgoingStock con picked_quantity se necessario
            if outgoing_item.quantity != line.picked_quantity:
                outgoing_item.quantity = line.picked_quantity
            db.delete(outgoing_item) # Rimuovi da outgoing stock
        
        if line.requested_quantity != line.picked_quantity:
            raise HTTPException(status_code=400, detail=f"Order line {line.id} not fully picked. Requested: {line.requested_quantity}, Picked: {line.picked_quantity}")

    order.is_completed = True
    
    # LOGGING: Registra l'evasione dell'ordine
    logger = LoggingService(db)
    
    # Prepara dettagli per il log
    order_summary = []
    total_items_fulfilled = 0
    
    for line in order.lines:
        order_summary.append({
            'product_sku': line.product_sku,
            'requested_quantity': line.requested_quantity,
            'picked_quantity': line.picked_quantity
        })
        total_items_fulfilled += line.picked_quantity
    
    # Logga ogni prodotto evaso separatamente per visibilità nelle colonne SKU/Ubicazioni  
    for line in order.lines:
        logger.log_operation(
            operation_type=OperationType.ORDINE_EVASO,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=line.product_sku,  # SKU del prodotto evaso
            location_from="OUTGOING",  # Da giacenza in uscita
            location_to=None,  # Evasione (esce dal magazzino)
            quantity=line.picked_quantity,  # Quantità evasa
            user_id="fulfill_user",  # TODO: Sostituire con sistema auth reale
            file_name=f"ORDER_{order.order_number}",  # Numero ordine nella colonna Dettagli
            details={
                'order_number': order.order_number,  # Numero ordine nei dettagli JSON
                'customer_name': order.customer_name,
                'order_date': order.order_date.isoformat() if order.order_date else None,
                'operation_description': f"Evasione ordine {order.order_number}: evaso {line.picked_quantity}x {line.product_sku} per cliente {order.customer_name}",
                'fulfill_type': 'manual_fulfill',
                'order_line_id': line.id,
                'requested_quantity': line.requested_quantity,
                'total_order_items': total_items_fulfilled
            },
            api_endpoint=f"/orders/{order_id}/fulfill"
        )
    
    db.commit()
    db.refresh(order)
    return order

# --- Nuova Funzionalità: Picking da File TXT ---

async def _parse_picking_file_scanner(file: UploadFile, db: Session) -> Tuple[Dict[str, Dict[str, Dict[str, int]]], List[Dict[str, Any]]]:
    """
    Parsing del file di picking da pistola scanner con formato:
    - Riga 1: Numero Ordine
    - Riga 2: Ubicazione
    - Righe 3+: EAN/SKU ripetuti (quantità = numero ripetizioni) oppure EAN/SKU_quantità
    - Il pattern si ripete per più ordini
    
    Ritorna: (parsed_data, parse_errors)
    """
    content = await file.read()
    try:
        text_content = content.decode("utf-8")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Il file non è in formato UTF-8 valido.")
    
    lines = [line.strip() for line in text_content.splitlines() if line.strip()]
    parsed_data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    parse_errors = []
    
    # Parsing migliorato per il formato scanner
    current_order = None
    current_location = None
    
    # Ottieni le ubicazioni esistenti dal database per confronto
    all_locations = {loc.name for loc in db.query(models.Location).all()}
    
    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        # Verifica se è un numero ordine (numerico o alfanumerico corto)
        if line.replace('-', '').replace('_', '').isalnum() and len(line) <= 10:
            # Verifica che non sia una ubicazione esistente
            if line not in all_locations:
                current_order = line
                current_location = None
                continue
        
        # Verifica se è una ubicazione esistente nel database
        if line in all_locations:
            current_location = line
            continue
        
        # Altrimenti è un prodotto
        if current_order and current_location:
            ean_or_sku = line
            quantity = 1
            
            # Parsing quantità da formato SKU_quantità
            if '_' in ean_or_sku:
                parts = ean_or_sku.rsplit('_', 1)
                if len(parts) == 2 and parts[1].isdigit():
                    ean_or_sku = parts[0]
                    quantity = int(parts[1])
            
            # Trova SKU nel database
            sku_found = None
            try:
                # Prova come EAN
                ean_code = db.query(models.EanCode).filter(models.EanCode.ean == ean_or_sku).first()
                if ean_code:
                    sku_found = ean_code.product_sku
                else:
                    # Prova come SKU diretto
                    product = db.query(models.Product).filter(models.Product.sku == ean_or_sku).first()
                    if product:
                        sku_found = product.sku
            except Exception:
                # Ignora errori di database per evitare hang
                pass
            
            if sku_found:
                parsed_data[current_order][current_location][sku_found] += quantity
            else:
                parse_errors.append({
                    "line": i+1,
                    "message": f"EAN/SKU '{ean_or_sku}' non trovato nel database",
                    "field": "sku",
                    "value": ean_or_sku,
                    "order": current_order,
                    "location": current_location
                })
        else:
            # Gestisci casi speciali quando mancano ordine o ubicazione
            if current_order and not current_location:
                # Prodotto senza ubicazione - deve richiedere ubicazione
                parse_errors.append({
                    "line": i+1,
                    "message": f"Prodotto '{line}' senza ubicazione (ordine: {current_order})",
                    "field": "location",
                    "value": line,
                    "order": current_order,
                    "location": "MANCANTE"
                })
            elif not current_order:
                # Prodotto senza ordine
                parse_errors.append({
                    "line": i+1,
                    "message": f"Prodotto '{line}' senza numero ordine",
                    "field": "order",
                    "value": line,
                    "order": "MANCANTE",
                    "location": current_location or "MANCANTE"
                })
            else:
                parse_errors.append({
                    "line": i+1,
                    "message": f"Prodotto '{line}' senza ordine o ubicazione validi",
                    "field": "general",
                    "value": line,
                    "order": current_order or "MANCANTE",
                    "location": current_location or "MANCANTE"
                })
        
    return parsed_data, parse_errors

@router.get("/{order_id}/picking-list-print")
async def get_picking_list_print(order_id: int, db: Session = Depends(get_db)):
    """
    Genera una versione stampabile della picking list per un ordine.
    """
    order = db.query(models.Order).options(joinedload(models.Order.lines)).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Usa lo stesso sistema di prenotazioni della UI per coerenza
    from wms_app.services.reservation_service import ReservationService
    
    reservation_service = ReservationService(db)
    reservation_service.cleanup_expired_reservations()
    
    # Prepara lista prodotti necessari per l'ordine (stesso formato della UI)
    products_needed = []
    for line in order.lines:
        if line.requested_quantity > line.picked_quantity:
            remaining_to_pick = line.requested_quantity - line.picked_quantity
            products_needed.append({
                'sku': line.product_sku,
                'quantity': remaining_to_pick,
                'line_id': line.id
            })
    
    suggestions = {}
    
    if products_needed:
        try:
            # Usa il sistema di prenotazioni per allocare ubicazioni (come nella UI)
            allocations = reservation_service.allocate_picking_locations(
                order_id=str(order.order_number), 
                products_needed=products_needed
            )
            
            # Converte il risultato nel formato per la stampa
            for allocation in allocations:
                sku = allocation['sku']
                
                product_suggestions = []
                for loc_allocation in allocation['allocations']:
                    product_suggestions.append({
                        "location_name": loc_allocation['location_name'],
                        "quantity": loc_allocation['quantity']
                    })
                
                suggestions[sku] = {
                    "needed": allocation['requested_quantity'],
                    "locations": product_suggestions
                }
                
        except Exception as e:
            # Fallback: usa inventario fisico se il sistema prenotazioni fallisce
            for line in order.lines:
                if line.requested_quantity <= line.picked_quantity:
                    continue
                    
                remaining_to_pick = line.requested_quantity - line.picked_quantity
                product_sku = line.product_sku
                
                available_stock = db.query(models.Inventory).filter(
                    models.Inventory.product_sku == product_sku,
                    models.Inventory.quantity > 0
                ).order_by(
                    models.Inventory.location_name
                ).all()
                
                product_suggestions = []
                for item in available_stock:
                    if remaining_to_pick <= 0:
                        break
                    
                    qty_from_location = min(remaining_to_pick, item.quantity)
                    product_suggestions.append({
                        "location_name": item.location_name,
                        "quantity": qty_from_location
                    })
                    remaining_to_pick -= qty_from_location
                
                suggestions[product_sku] = {
                    "needed": line.requested_quantity - line.picked_quantity,
                    "locations": product_suggestions
                }
    
    # Genera HTML stampabile
    # Prepara il numero ordine per il JavaScript
    order_number = order.order_number
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Picking List - Ordine {order_number}</title>
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Libre+Barcode+39&display=swap" rel="stylesheet">
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .header {{ border-bottom: 2px solid #000; padding-bottom: 10px; margin-bottom: 20px; }}
            
            /* Layout a tre colonne per header */
            .header-content {{ 
                display: flex; 
                justify-content: space-between; 
                align-items: flex-start;
                margin-bottom: 20px;
            }}
            
            .order-info {{ 
                flex: 1; 
                padding-right: 20px;
            }}
            
            .barcode-container {{ 
                flex: 1; 
                text-align: center; 
                padding: 0 20px;
            }}
            .barcode {{ font-family: 'Libre Barcode 39', monospace; font-size: 48px; margin: 10px 0; }}
            .barcode-text {{ font-size: 14px; font-weight: bold; margin-top: 5px; }}
            
            .products-recap {{ 
                flex: 1; 
                padding-left: 20px;
                border-left: 1px solid #ccc;
            }}
            .products-recap h3 {{ 
                margin: 0 0 10px 0; 
                font-size: 16px; 
                color: #333;
            }}
            .product-item {{ 
                margin: 5px 0; 
                font-size: 14px; 
                display: flex; 
                justify-content: space-between;
                color: #000;
            }}
            .product-sku {{ 
                font-weight: bold; 
                color: #000;
            }}
            .product-qty {{ 
                font-weight: bold;
                color: #000;
            }}
            
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #000; padding: 8px; text-align: left; }}
            th {{ background-color: #f0f0f0; }}
            .location {{ font-weight: bold; }}
            @media print {{ 
                body {{ margin: 0; }}
                .no-print {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>PICKING LIST</h1>
        </div>
        
        <div class="header-content">
            <!-- Colonna Sinistra: Info Ordine -->
            <div class="order-info">
                <p><strong>Numero Ordine:</strong> {order_number}</p>
                <p><strong>Cliente:</strong> {order.customer_name}</p>
                <p><strong>Data:</strong> {order.order_date.strftime('%d/%m/%Y %H:%M') if order.order_date else 'N/A'}</p>
            </div>
            
            <!-- Colonna Centrale: Barcode -->
            <div class="barcode-container">
                <div class="barcode">*{order_number}*</div>
                <div class="barcode-text">{order_number}</div>
            </div>
            
            <!-- Colonna Destra: Recap Prodotti -->
            <div class="products-recap">
                <h3>Quantità Ordine:</h3>"""
    
    # Aggiungi il recap dei prodotti richiesti
    for line in order.lines:
        remaining_to_pick = line.requested_quantity - line.picked_quantity
        if remaining_to_pick > 0:
            html_content += f"""
                <div class="product-item">
                    <span class="product-sku">{line.product_sku}</span>
                    <span class="product-qty">{remaining_to_pick} pz</span>
                </div>"""
    
    html_content += """
            </div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>SKU</th>
                    <th>Ubicazione</th>
                    <th>Quantità da Prelevare</th>
                    <th>☐ Prelevato</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for sku, suggestion in suggestions.items():
        for location in suggestion["locations"]:
            html_content += f"""
                <tr>
                    <td>{sku}</td>
                    <td class="location">{location['location_name']}</td>
                    <td>{location['quantity']}</td>
                    <td style="text-align: center; width: 50px;">☐</td>
                </tr>
            """
    
    html_content += """
            </tbody>
        </table>
        <div class="no-print">
            <button onclick="window.print()">Stampa</button>
            <button onclick="window.close()">Chiudi</button>
        </div>
        
        <script>
            // Auto-print quando la pagina è caricata
            window.onload = function() {{
                setTimeout(function() {{
                    window.print();
                }}, 500);
            }}
        </script>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)

# --- Nuovi Endpoint per Gestione Archiviazione e Annullamento ---

@router.post("/{order_id}/archive")
def archive_order(order_id: int, fulfillment_request: schemas.FulfillmentRequest, db: Session = Depends(get_db)):
    """Archivia un ordine completato o annullato."""
    try:
        order = db.query(models.Order).filter(models.Order.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Validazione stato ordine
        
        if not order.is_completed and not order.is_cancelled:
            raise HTTPException(status_code=400, detail="Can only archive completed or cancelled orders")
        
        if order.is_archived:
            raise HTTPException(status_code=400, detail="Order is already archived")
        
        # Archivia l'ordine
        order.is_archived = True
        order.archived_date = datetime.utcnow()
        
        # Salva il numero DDT se fornito
        if fulfillment_request.ddt_number:
            order.ddt_number = fulfillment_request.ddt_number.strip()
        
        db.commit()
        return {"message": f"Order {order.order_number} archived successfully"}
        
    except HTTPException:
        # Re-raise HTTPExceptions as-is
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error archiving order: {str(e)}")

@router.post("/{order_id}/cancel")
def cancel_order(order_id: int, db: Session = Depends(get_db)):
    """Annulla un ordine e rilascia la giacenza in uscita."""
    try:
        order = db.query(models.Order).filter(models.Order.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Validazione stato ordine
        
        if order.is_completed:
            raise HTTPException(status_code=400, detail="Cannot cancel a completed order")
        
        if order.is_cancelled:
            raise HTTPException(status_code=400, detail="Order is already cancelled")
        
        if order.is_archived:
            raise HTTPException(status_code=400, detail="Cannot cancel an archived order")
        
        # Recupera tutte le giacenze in uscita per questo ordine
        outgoing_stocks = db.query(models.OutgoingStock).join(models.OrderLine).filter(
            models.OrderLine.order_id == order_id
        ).all()
        
        released_items = []
        inventory_restored = []
        
        for stock in outgoing_stocks:
            released_items.append({
                "product_sku": stock.product_sku,
                "quantity": stock.quantity
            })
            
            # NUOVO: Ripristina giacenza in inventario (ipotesi: location TERRA)
            # Cerca se esiste già giacenza per questo prodotto in TERRA
            terra_inventory = db.query(models.Inventory).filter(
                models.Inventory.product_sku == stock.product_sku,
                models.Inventory.location_name == "TERRA"
            ).first()
            
            if terra_inventory:
                terra_inventory.quantity += stock.quantity
            else:
                # Crea nuovo record inventario in TERRA
                new_inventory = models.Inventory(
                    product_sku=stock.product_sku,
                    location_name="TERRA",
                    quantity=stock.quantity
                )
                db.add(new_inventory)
            
            inventory_restored.append({
                "product_sku": stock.product_sku,
                "quantity": stock.quantity,
                "restored_to": "TERRA"
            })
            
            # Rimuovi dalla giacenza in uscita
            db.delete(stock)
        
        # Annulla l'ordine
        order.is_cancelled = True
        order.cancelled_date = datetime.utcnow()
        
        # LOGGING: Registra l'annullamento dell'ordine
        logger = LoggingService(db)
        
        # Prepara dettagli prodotti per log leggibile
        restored_products_summary = []
        for item in inventory_restored:
            restored_products_summary.append(f"{item['quantity']}x {item['product_sku']} → {item['restored_to']}")
        
        # Logga ogni prodotto ripristinato separatamente per visibilità nelle colonne SKU/Ubicazioni
        for item in inventory_restored:
            logger.log_operation(
                operation_type=OperationType.ORDINE_ANNULLATO,
                operation_category=OperationCategory.MANUAL,
                status=OperationStatus.SUCCESS,
                product_sku=item['product_sku'],  # SKU del prodotto ripristinato
                location_from=None,  # Da OutgoingStock (non ha ubicazione fisica)
                location_to=item['restored_to'],  # Ripristinato a TERRA
                quantity=item['quantity'],  # Quantità ripristinata
                user_id="cancel_user",  # TODO: Sostituire con sistema auth reale
                file_name=f"ORDER_{order.order_number}",  # Numero ordine nella colonna Dettagli
                details={
                    'order_number': order.order_number,  # Numero ordine nei dettagli JSON
                    'customer_name': order.customer_name,
                    'cancelled_date': order.cancelled_date.isoformat(),
                    'operation_description': f"Annullamento ordine {order.order_number}: ripristinato {item['quantity']}x {item['product_sku']} in {item['restored_to']} per riposizionamento",
                    'cancel_type': 'manual_cancel',
                    'total_order_items': sum(restored['quantity'] for restored in inventory_restored),
                    'restoration_reason': 'order_cancellation'
                },
                api_endpoint=f"/orders/{order_id}/cancel"
            )
        
        db.commit()
        
        return {
            "message": f"Order {order.order_number} cancelled successfully",
            "released_items": released_items,
            "inventory_restored": inventory_restored,
            "note": "I prodotti sono stati automaticamente ripristinati in ubicazione TERRA in attesa di riposizionamento manuale da parte degli operatori"
        }
    except HTTPException:
        # Re-raise HTTPExceptions as-is
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error cancelling order: {str(e)}")

@router.delete("/{order_id}/unarchive")
def unarchive_order(order_id: int, db: Session = Depends(get_db)):
    """Rimuove un ordine dall'archivio (riporta nella lista normale)."""
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if not order.is_archived:
        raise HTTPException(status_code=400, detail="Order is not archived")
    
    # Rimuovi dall'archivio
    order.is_archived = False
    order.archived_date = None
    
    try:
        db.commit()
        return {"message": f"Order {order.order_number} removed from archive successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error unarchiving order: {str(e)}")

# --- Sistema Import Automatico da Cartella ---

@router.post("/auto-import/configure-folder")
def configure_auto_import_folder(request: dict, db: Session = Depends(get_db)):
    """Configura il percorso della cartella per l'import automatico."""
    try:
        folder_path = request.get("folder_path")
        if not folder_path:
            raise HTTPException(status_code=400, detail="folder_path è richiesto")
        
        # Valida che la cartella esista
        if not os.path.exists(folder_path):
            raise HTTPException(status_code=400, detail=f"La cartella {folder_path} non esiste")
        
        if not os.path.isdir(folder_path):
            raise HTTPException(status_code=400, detail=f"Il percorso {folder_path} non è una cartella")
        
        # Crea le sottocartelle se non esistono
        processed_folder = os.path.join(folder_path, "processati")
        error_folder = os.path.join(folder_path, "errori")
        
        os.makedirs(processed_folder, exist_ok=True)
        os.makedirs(error_folder, exist_ok=True)
        
        # Salva la configurazione nel database
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == "auto_import_folder").first()
        if setting:
            setting.value = folder_path
            setting.updated_at = datetime.utcnow()
        else:
            setting = models.SystemSetting(
                key="auto_import_folder",
                value=folder_path,
                description="Cartella monitorata per import automatico ordini"
            )
            db.add(setting)
        
        db.commit()
        
        return {
            "message": "Cartella configurata con successo",
            "folder_path": folder_path,
            "processed_folder": processed_folder,
            "error_folder": error_folder
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Errore nella configurazione: {str(e)}")

@router.get("/auto-import/folder-config")
def get_auto_import_config(db: Session = Depends(get_db)):
    """Recupera la configurazione attuale della cartella di import."""
    setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == "auto_import_folder").first()
    
    if not setting:
        return {"configured": False, "folder_path": None}
    
    folder_exists = os.path.exists(setting.value) and os.path.isdir(setting.value)
    
    return {
        "configured": True,
        "folder_path": setting.value,
        "folder_exists": folder_exists,
        "last_updated": setting.updated_at.isoformat() if setting.updated_at else None
    }

@router.post("/auto-import/from-folder")
def auto_import_from_folder(db: Session = Depends(get_db)):
    """Esegue l'import automatico di tutti i file dalla cartella configurata."""
    try:
        # Recupera la configurazione della cartella
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == "auto_import_folder").first()
        if not setting:
            raise HTTPException(status_code=400, detail="Cartella di import non configurata")
        
        folder_path = setting.value
        if not os.path.exists(folder_path):
            raise HTTPException(status_code=400, detail=f"Cartella {folder_path} non trovata")
        
        processed_folder = os.path.join(folder_path, "processati")
        error_folder = os.path.join(folder_path, "errori")
        
        # Assicurati che le cartelle esistano
        os.makedirs(processed_folder, exist_ok=True)
        os.makedirs(error_folder, exist_ok=True)
        
        # Cerca file TXT e CSV nella cartella
        supported_extensions = ['.txt', '.csv']
        files_found = []
        
        for file_path in Path(folder_path).iterdir():
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                files_found.append(file_path)
        
        if not files_found:
            return {
                "message": "Nessun file da processare",
                "files_processed": 0,
                "files_with_errors": 0,
                "details": []
            }
        
        # Processa ogni file
        results = []
        files_processed = 0
        files_with_errors = 0
        
        for file_path in files_found:
            try:
                # Leggi il file con gestione BOM e caratteri invisibili
                file_content = _read_file_with_bom_handling(file_path)
                
                # Processa il file usando la logica esistente
                result = _process_orders_file_content(file_content, str(file_path), db)
                
                if result["success"]:
                    # Sposta il file nella cartella processati
                    destination = os.path.join(processed_folder, file_path.name)
                    shutil.move(str(file_path), destination)
                    files_processed += 1
                    results.append({
                        "file": file_path.name,
                        "status": "processed",
                        "message": result["message"],
                        "orders_created": result.get("orders_created", 0),
                        "orders_details": result.get("orders_details", []),
                        "general_errors": result.get("general_errors", [])
                    })
                else:
                    # Sposta il file nella cartella errori
                    destination = os.path.join(error_folder, file_path.name)
                    shutil.move(str(file_path), destination)
                    files_with_errors += 1
                    results.append({
                        "file": file_path.name,
                        "status": "error",
                        "message": result["message"],
                        "errors": result.get("errors", [])
                    })
                    
            except Exception as file_error:
                # Sposta il file nella cartella errori
                try:
                    destination = os.path.join(error_folder, file_path.name)
                    shutil.move(str(file_path), destination)
                except:
                    pass  # Se non riesce a spostare, continua
                
                files_with_errors += 1
                results.append({
                    "file": file_path.name,
                    "status": "error",
                    "message": f"Errore nel processamento del file: {str(file_error)}",
                    "errors": [str(file_error)]
                })
        
        return {
            "message": f"Import completato: {files_processed} file processati, {files_with_errors} errori",
            "files_processed": files_processed,
            "files_with_errors": files_with_errors,
            "details": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante l'import automatico: {str(e)}")

# --- Funzioni di supporto per import automatico ---

def _read_file_with_bom_handling(file_path: Path) -> str:
    """Legge un file gestendo BOM e caratteri invisibili."""
    encodings_to_try = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    
    for encoding in encodings_to_try:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            
            # Rimuovi caratteri invisibili comuni
            content = content.replace('\ufeff', '')  # BOM UTF-8
            content = content.replace('\u200b', '')  # Zero-width space
            content = content.replace('\u00a0', ' ')  # Non-breaking space
            
            return content.strip()
            
        except UnicodeDecodeError:
            continue
    
    raise ValueError(f"Impossibile leggere il file {file_path.name} con nessuna codifica supportata")

def _process_orders_file_content(content: str, filename: str, db: Session) -> dict:
    """Processa il contenuto di un file ordini e crea gli ordini nel database."""
    try:
        lines = [line.strip() for line in content.split('\n') if line.strip()]
        
        if not lines:
            return {"success": False, "message": "File vuoto", "errors": ["Il file non contiene dati"]}
        
        orders_data = defaultdict(lambda: {"customer_name": "", "lines": [], "warnings": []})
        errors = []
        line_number = 0
        
        for line in lines:
            line_number += 1
            
            # Supporta sia CSV che formato separato da virgole
            parts = [part.strip() for part in line.split(',')]
            
            if len(parts) < 4:
                errors.append(f"Riga {line_number}: formato non valido (necessari almeno 4 campi)")
                continue
            
            order_number, customer_name, product_sku, quantity_str = parts[:4]
            
            try:
                quantity = int(quantity_str)
                if quantity <= 0:
                    errors.append(f"Riga {line_number}: quantità deve essere maggiore di 0")
                    continue
            except ValueError:
                errors.append(f"Riga {line_number}: quantità non valida '{quantity_str}'")
                continue
            
            # Verifica che il prodotto esista
            product = db.query(models.Product).filter(models.Product.sku == product_sku).first()
            if not product:
                orders_data[order_number]["warnings"].append(f"Prodotto '{product_sku}' non trovato in anagrafica")
                errors.append(f"Riga {line_number}: prodotto '{product_sku}' non trovato")
                continue
            
            # Aggiungi ai dati dell'ordine
            orders_data[order_number]["customer_name"] = customer_name
            orders_data[order_number]["lines"].append({
                "product_sku": product_sku,
                "requested_quantity": quantity
            })
        
        if errors and len(orders_data) == 0:
            return {"success": False, "message": "Nessun ordine valido trovato", "errors": errors}
        
        # Crea gli ordini nel database
        orders_created = 0
        
        for order_number, order_data in orders_data.items():
            try:
                # Verifica se l'ordine esiste già
                existing_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
                if existing_order:
                    orders_data[order_number]["warnings"].append(f"Ordine già esistente nel database")
                    errors.append(f"Ordine {order_number} già esistente")
                    continue
                
                # Crea il nuovo ordine
                new_order = models.Order(
                    order_number=order_number,
                    customer_name=order_data["customer_name"],
                    order_date=datetime.utcnow()
                )
                db.add(new_order)
                db.flush()  # Per ottenere l'ID
                
                # Crea le righe d'ordine
                for line_data in order_data["lines"]:
                    order_line = models.OrderLine(
                        order_id=new_order.id,
                        product_sku=line_data["product_sku"],
                        requested_quantity=line_data["requested_quantity"]
                    )
                    db.add(order_line)
                
                orders_created += 1
                
            except Exception as order_error:
                errors.append(f"Errore creazione ordine {order_number}: {str(order_error)}")
                db.rollback()
                continue
        
        # LOGGING: Registra l'import automatico da cartella
        if orders_created > 0:
            logger = LoggingService(db)
            
            # Logga ogni prodotto di ogni ordine creato automaticamente
            for order_number, order_data in orders_data.items():
                created_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
                if created_order:  # Solo se l'ordine è stato creato con successo
                    for line_data in order_data["lines"]:
                        logger.log_operation(
                            operation_type=OperationType.ORDINE_CREATO,
                            operation_category=OperationCategory.SYSTEM,  # Automatico
                            status=OperationStatus.SUCCESS,
                            product_sku=line_data["product_sku"],  # SKU del prodotto importato
                            quantity=line_data["requested_quantity"],  # Quantità richiesta
                            user_id="auto_import_system",
                            file_name=f"ORDER_{order_number}",  # Numero ordine nella colonna Dettagli
                            details={
                                'order_number': order_number,  # Numero ordine nei dettagli JSON
                                'customer_name': order_data["customer_name"],
                                'creation_method': 'auto_import',
                                'source_file': filename,
                                'operation_description': f"Import automatico: creato ordine {order_number}, aggiunto {line_data['requested_quantity']}x {line_data['product_sku']} per cliente {order_data['customer_name']}",
                                'auto_import_stats': {
                                    'orders_created': orders_created,
                                    'source_filename': filename
                                }
                            },
                            api_endpoint="/orders/auto-import/from-folder"
                        )
        
        if orders_created > 0:
            db.commit()
            message = f"Import completato: {orders_created} ordini creati da {filename}"
            if errors:
                message += f" (con {len(errors)} avvisi)"
            
            # Crea lista dettagliata degli ordini creati con i loro avvisi
            orders_details = []
            for order_number, order_data in orders_data.items():
                if order_number in [order_number for order_number, order_data in orders_data.items() if len(order_data.get("lines", [])) > 0]:
                    # Controlla se questo ordine è stato effettivamente creato
                    created_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
                    if created_order:
                        orders_details.append({
                            "order_number": order_number,
                            "customer_name": order_data["customer_name"],
                            "products_count": len(order_data["lines"]),
                            "warnings": order_data.get("warnings", [])
                        })
            
            return {
                "success": True, 
                "message": f"{orders_created} ordini creati",
                "orders_created": orders_created,
                "orders_details": orders_details,
                "general_errors": errors
            }
        else:
            db.rollback()
            return {"success": False, "message": "Nessun ordine creato", "errors": errors}
            
    except Exception as e:
        db.rollback()
        return {"success": False, "message": f"Errore durante il processamento: {str(e)}", "errors": [str(e)]}

# --- Endpoint per Cancellazione Completa Ordini ---

@router.delete("/{order_id}/delete")
def delete_order_completely(order_id: int, db: Session = Depends(get_db)):
    """Cancella completamente un ordine che non ha ancora iniziato il picking."""
    try:
        order = db.query(models.Order).filter(models.Order.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Verifica che l'ordine non abbia picking iniziato
        has_picking = False
        for line in order.lines:
            if line.picked_quantity > 0:
                has_picking = True
                break
        
        if has_picking:
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete order: picking has already started"
            )
        
        if order.is_completed:
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete order: order is already completed"
            )
        
        if order.is_archived:
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete order: order is archived"
            )
        
        # Verifica se ci sono prenotazioni attive (con import diretto)
        try:
            from wms_app.models.reservations import InventoryReservation
            active_reservations = db.query(InventoryReservation).filter(
                InventoryReservation.order_id == order.order_number,
                InventoryReservation.status == "active"
            ).count()
            
            if active_reservations > 0:
                raise HTTPException(
                    status_code=400,
                    detail="Cannot delete order: active reservations exist. Cancel the order first to release reservations."
                )
        except ImportError:
            # Se il modello non esiste, continua
            pass
        
        # Verifica se ci sono OutgoingStock associati
        try:
            outgoing_stocks = db.query(models.OutgoingStock).join(models.OrderLine).filter(
                models.OrderLine.order_id == order.id
            ).count()
            
            if outgoing_stocks > 0:
                raise HTTPException(
                    status_code=400,
                    detail="Cannot delete order: outgoing stock exists. Cancel the order first to release stock."
                )
        except Exception:
            # Se ci sono problemi con OutgoingStock, continua
            pass
        
        order_number = order.order_number
        
        # Elimina le righe dell'ordine
        db.query(models.OrderLine).filter(models.OrderLine.order_id == order.id).delete()
        
        # Elimina l'ordine
        db.delete(order)
        
        db.commit()
        
        return {
            "message": f"Order {order_number} deleted successfully",
            "order_number": order_number,
            "order_id": order_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting order: {str(e)}")

# === ENDPOINT PER PICKING IN TEMPO REALE ===

@router.post("/real-time-picking/scan-product")
async def scan_product_real_time(
    request_data: dict,
    db: Session = Depends(get_db)
):
    """
    Endpoint per processare la scansione di un prodotto durante il picking in tempo reale.
    Valida l'EAN code, verifica la corrispondenza con lo SKU richiesto e scala le giacenze.
    """
    try:
        order_id = request_data.get("order_id")
        location_name = request_data.get("location_name", "").upper()
        scanned_code = request_data.get("scanned_code", "")
        expected_sku = request_data.get("expected_sku", "")
        quantity = request_data.get("quantity", 1)
        
        if not all([order_id, location_name, scanned_code, expected_sku]):
            raise HTTPException(status_code=400, detail="Missing required parameters")
        
        # 1. Verifica che l'ordine esista e non sia completato
        order = db.query(models.Order).filter(
            models.Order.id == order_id,
            models.Order.is_completed == False
        ).first()
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found or already completed")
        
        # 2. Trova la riga dell'ordine corrispondente
        order_line = db.query(models.OrderLine).filter(
            models.OrderLine.order_id == order_id,
            models.OrderLine.product_sku == expected_sku
        ).first()
        
        if not order_line:
            # Log errore prodotto non nell'ordine
            logger = LoggingService(db)
            logger.log_error(
                operation_type=OperationType.PRELIEVO_TEMPO_REALE,
                error=f"Product {expected_sku} not found in order {order.order_number}",
                operation_category=OperationCategory.PICKING,
                product_sku=expected_sku,
                file_name=f"ORDER_{order.order_number}",
                details={
                    'order_number': order.order_number,
                    'expected_sku': expected_sku,
                    'scanned_code': scanned_code,
                    'location_name': location_name,
                    'error_reason': 'product_not_in_order',
                    'operation_description': f"Errore picking tempo reale ordine {order.order_number}: prodotto {expected_sku} non presente nell'ordine"
                },
                api_endpoint="/orders/real-time-picking/scan-product"
            )
            return {
                "success": False,
                "message": f"Prodotto {expected_sku} non trovato in questo ordine"
            }
        
        # 3. Controlla se c'è ancora quantità da prelevare
        remaining_to_pick = order_line.requested_quantity - order_line.picked_quantity
        if remaining_to_pick <= 0:
            return {
                "success": False,
                "message": f"Prodotto {expected_sku} già completamente prelevato"
            }
        
        # 4. Validazione barcode: controlla se è un EAN code o direttamente lo SKU
        product_sku = None
        
        # Prima prova a vedere se il codice scansionato è direttamente lo SKU
        product = db.query(models.Product).filter(models.Product.sku == scanned_code).first()
        if product:
            product_sku = product.sku
        else:
            # Altrimenti cerca negli EAN codes
            ean_code = db.query(models.EanCode).filter(models.EanCode.ean == scanned_code).first()
            if ean_code:
                product_sku = ean_code.product_sku
        
        if not product_sku:
            # Log errore barcode non riconosciuto
            logger = LoggingService(db)
            logger.log_error(
                operation_type=OperationType.PRELIEVO_TEMPO_REALE,
                error=f"Barcode '{scanned_code}' not recognized",
                operation_category=OperationCategory.PICKING,
                product_sku=expected_sku,
                location_from=location_name,
                file_name=f"ORDER_{order.order_number}",
                details={
                    'order_number': order.order_number,
                    'expected_sku': expected_sku,
                    'scanned_code': scanned_code,
                    'location_name': location_name,
                    'error_reason': 'barcode_not_recognized',
                    'operation_description': f"Errore picking tempo reale ordine {order.order_number}: barcode '{scanned_code}' non riconosciuto"
                },
                api_endpoint="/orders/real-time-picking/scan-product"
            )
            return {
                "success": False,
                "message": f"Codice scansionato '{scanned_code}' non riconosciuto"
            }
        
        # 5. Verifica che il prodotto scansionato corrisponda a quello richiesto
        if product_sku != expected_sku:
            return {
                "success": False,
                "message": f"Prodotto errato! Richiesto: {expected_sku}, Scansionato: {product_sku}"
            }
        
        # 6. Verifica disponibilità nella specifica ubicazione
        inventory_item = db.query(models.Inventory).filter(
            models.Inventory.product_sku == product_sku,
            models.Inventory.location_name == location_name,
            models.Inventory.quantity > 0
        ).first()
        
        if not inventory_item:
            return {
                "success": False,
                "message": f"Prodotto {product_sku} non disponibile nell'ubicazione {location_name}"
            }
        
        # 7. Determina la quantità effettiva da prelevare
        actual_quantity = min(quantity, inventory_item.quantity, remaining_to_pick)
        
        # 8. Scala la giacenza in tempo reale
        inventory_item.quantity -= actual_quantity
        
        # 9. Se l'ubicazione rimane vuota, elimina il record di inventario
        if inventory_item.quantity <= 0:
            db.delete(inventory_item)
        
        # 10. Aggiorna la quantità prelevata nell'ordine
        order_line.picked_quantity += actual_quantity
        
        # 11. Sposta in OutgoingStock (come nel picking manuale)
        outgoing_item = db.query(models.OutgoingStock).filter(
            models.OutgoingStock.order_line_id == order_line.id,
            models.OutgoingStock.product_sku == product_sku
        ).first()

        if outgoing_item:
            outgoing_item.quantity += actual_quantity
        else:
            new_outgoing_item = models.OutgoingStock(
                order_line_id=order_line.id,
                product_sku=product_sku,
                quantity=actual_quantity
            )
            db.add(new_outgoing_item)
        
        # 12. LOGGING: Registra l'operazione di picking in tempo reale
        logger = LoggingService(db)
        logger.log_operation(
            operation_type=OperationType.PRELIEVO_TEMPO_REALE,
            operation_category=OperationCategory.PICKING,
            status=OperationStatus.SUCCESS,
            product_sku=product_sku,
            location_from=location_name,
            location_to=None,  # Picking: scala da inventario
            quantity=actual_quantity,
            user_id="realtime_picker",  # TODO: Sostituire con sistema auth reale
            details={
                'order_number': order.order_number,
                'order_id': order_id,
                'scanned_code': scanned_code,
                'expected_sku': expected_sku,
                'operation_description': f"Picking tempo reale: {product_sku} ({actual_quantity} pz) da {location_name} per ordine {order.order_number}",
                'picking_type': 'real_time_picking',
                'barcode_validation': 'passed',
                'customer_name': order.customer_name,
                'remaining_in_location': inventory_item.quantity if inventory_item.quantity > 0 else 0,
                'remaining_to_pick': order_line.requested_quantity - order_line.picked_quantity - actual_quantity
            },
            api_endpoint="/orders/real-time-picking/scan-product"
        )
        
        # 13. Commit delle modifiche
        db.commit()
        
        # 14. Prepara la risposta di successo
        return {
            "success": True,
            "message": f"Prodotto prelevato con successo",
            "product_sku": product_sku,
            "location_name": location_name,
            "quantity_picked": actual_quantity,
            "remaining_in_location": inventory_item.quantity if inventory_item.quantity > 0 else 0,
            "remaining_to_pick": order_line.requested_quantity - order_line.picked_quantity,
            "order_line_completed": (order_line.requested_quantity - order_line.picked_quantity) <= 0
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing real-time picking: {str(e)}")

# === ENDPOINT IMPORT ORDINI DA EXCEL ===

@router.post("/parse-excel-orders")
async def parse_excel_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Parsing e validazione di ordini da file Excel (.xlsx).
    Formato: Colonna B = N ordine, T = Cliente, Q = SKU, S = Quantità
    
    Ritorna un recap dettagliato per conferma dell'utente prima del commit.
    """
    try:
        # Verifica che sia un file Excel
        if not file.filename.lower().endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Solo file Excel (.xlsx) sono supportati")
        
        # Leggi il contenuto del file
        content = await file.read()
        
        # Importa openpyxl dinamicamente
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError as e:
            import sys
            raise HTTPException(status_code=500, detail=f"openpyxl non disponibile. Python: {sys.executable}, Error: {str(e)}")
        
        # Carica il workbook Excel
        try:
            workbook = load_workbook(BytesIO(content), data_only=True)
            worksheet = workbook.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Errore lettura file Excel: {str(e)}")
        
        # Parsing dei dati dalle colonne identificate
        orders_data = defaultdict(lambda: {"customer_name": "", "lines": defaultdict(int), "errors": [], "warnings": []})
        parse_errors = []
        line_number = 1  # Riga 1 = header
        
        # Leggi gli headers per identificare le colonne corrette
        headers = []
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip() if cell is not None else '' for cell in row]
            break
        
        # Trova gli indici delle colonne che ci interessano
        column_mapping = {}
        target_headers = {
            'order_number': 'CODICE ORDINE MASTER',
            'customer_name': 'RAGIONE SOCIALE DESTINATARIO', 
            'sku': 'CODICE PADRE PRODOTTO',
            'quantity': 'Q PRODOTTO'
        }
        
        for key, target_header in target_headers.items():
            try:
                index = headers.index(target_header)
                column_mapping[key] = index
# Debug rimosso per produzione
            except ValueError:
                parse_errors.append({
                    "line": 1,
                    "message": f"Colonna '{target_header}' non trovata negli headers",
                    "field": "excel_headers",
                    "value": f"Headers disponibili: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}"
                })
        
        # Verifica che tutte le colonne siano state trovate
        if len(column_mapping) != 4:
            missing = set(target_headers.keys()) - set(column_mapping.keys())
            raise HTTPException(
                status_code=400, 
                detail=f"Colonne mancanti nel file Excel: {', '.join(target_headers[k] for k in missing)}"
            )
        
        # Itera attraverso tutte le righe (saltando la prima che è l'header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            line_number += 1
            
            try:
                # Estrae valori dalle colonne identificate dinamicamente
                order_number = row[column_mapping['order_number']] if len(row) > column_mapping['order_number'] and row[column_mapping['order_number']] is not None else None
                sku = row[column_mapping['sku']] if len(row) > column_mapping['sku'] and row[column_mapping['sku']] is not None else None
                quantity_raw = row[column_mapping['quantity']] if len(row) > column_mapping['quantity'] and row[column_mapping['quantity']] is not None else None
                customer_name = row[column_mapping['customer_name']] if len(row) > column_mapping['customer_name'] and row[column_mapping['customer_name']] is not None else None
                
            except Exception as e:
                parse_errors.append({
                    "line": line_number,
                    "message": f"Errore lettura riga Excel: {str(e)}",
                    "field": "excel_structure",
                    "value": f"Row length: {len(row) if row else 'None'}"
                })
                continue
            
            # Salta righe completamente vuote
            if all(v is None or str(v).strip() == '' for v in [order_number, sku, quantity_raw, customer_name]):
                continue
            
            # Validazioni base
            validation_errors = []
            
            if not order_number or str(order_number).strip() == '':
                validation_errors.append(f"Numero ordine mancante (colonna B)")
            else:
                order_number = str(order_number).strip()
            
            if not customer_name or str(customer_name).strip() == '':
                validation_errors.append(f"Nome cliente mancante (colonna T)")
            else:
                customer_name = str(customer_name).strip()
            
            if not sku or str(sku).strip() == '':
                validation_errors.append(f"SKU prodotto mancante (colonna Q)")
            else:
                sku = str(sku).strip()
            
            # Validazione quantità
            try:
                if quantity_raw is None:
                    validation_errors.append(f"Quantità mancante (colonna S)")
                    quantity = 0
                else:
                    quantity = int(float(quantity_raw))  # Gestisce sia int che float da Excel
                    if quantity <= 0:
                        validation_errors.append(f"Quantità deve essere maggiore di 0")
            except (ValueError, TypeError):
                validation_errors.append(f"Quantità non valida: '{quantity_raw}' (colonna S)")
                quantity = 0
            
            # Se ci sono errori di validazione, registra e continua
            if validation_errors:
                parse_errors.extend([{
                    "line": line_number,
                    "message": error,
                    "field": "validation",
                    "value": f"B:{order_number}, Q:{sku}, S:{quantity_raw}, T:{customer_name}"
                } for error in validation_errors])
                continue
            
            # Verifica che il prodotto esista nel database
            product = db.query(models.Product).filter(models.Product.sku == sku).first()
            if not product:
                orders_data[order_number]["errors"].append(f"Riga {line_number}: SKU '{sku}' non trovato in anagrafica")
                parse_errors.append({
                    "line": line_number,
                    "message": f"SKU '{sku}' non trovato in anagrafica",
                    "field": "sku",
                    "value": sku
                })
                continue
            
            # Verifica ordine duplicato esistente nel database
            existing_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
            if existing_order:
                orders_data[order_number]["warnings"].append(f"Ordine '{order_number}' già esistente nel database")
            
            # Aggiungi/aggiorna i dati dell'ordine
            orders_data[order_number]["customer_name"] = customer_name
            
            # Consolidamento automatico per SKU duplicati
            if sku in orders_data[order_number]["lines"]:
                old_quantity = orders_data[order_number]["lines"][sku]
                orders_data[order_number]["lines"][sku] += quantity
                orders_data[order_number]["warnings"].append(
                    f"SKU '{sku}' consolidato: {old_quantity} + {quantity} = {orders_data[order_number]['lines'][sku]}"
                )
            else:
                orders_data[order_number]["lines"][sku] = quantity
        
        # Prepara recap dettagliato
        recap_items = []
        total_orders = len(orders_data)
        total_lines = 0
        errors_count = len(parse_errors)
        warnings_count = 0
        
        # Genera recap items per ogni ordine e ogni prodotto
        for order_number, order_data in orders_data.items():
            warnings_count += len(order_data["warnings"])
            for sku, quantity in order_data["lines"].items():
                total_lines += 1
                
                # Trova il prodotto per descrizione
                product = db.query(models.Product).filter(models.Product.sku == sku).first()
                description = product.description if product else "Prodotto non trovato"
                
                # Determina lo stato
                status = "ok"
                if order_data["errors"]:
                    status = "error"
                elif order_data["warnings"]:
                    status = "warning"
                
                recap_items.append({
                    "line": len(recap_items) + 1,
                    "order_number": order_number,
                    "customer_name": order_data["customer_name"],
                    "sku": sku,
                    "description": description,
                    "quantity": quantity,
                    "status": status,
                    "consolidation_applied": any("consolidato" in w for w in order_data["warnings"])
                })
        
        # Crea gli ordini direttamente (senza recap)
        orders_created = 0
        orders_updated = 0
        logger = LoggingService(db)
        
        for order_number, order_data in orders_data.items():
            if order_data["errors"]:  # Salta ordini con errori
                continue
                
            # Controlla se l'ordine esiste già
            existing_order = db.query(models.Order).filter(models.Order.order_number == str(order_number)).first()
            is_new_order = existing_order is None
            
            if existing_order and not existing_order.is_completed:
                # Aggiorna ordine esistente
                for sku, quantity in order_data["lines"].items():
                    # Controlla se la riga prodotto esiste già
                    existing_line = db.query(models.OrderLine).filter(
                        models.OrderLine.order_id == existing_order.id,
                        models.OrderLine.product_sku == sku
                    ).first()
                    
                    if existing_line:
                        existing_line.requested_quantity += quantity
                    else:
                        new_line = models.OrderLine(
                            order_id=existing_order.id,
                            product_sku=sku,
                            requested_quantity=quantity
                        )
                        db.add(new_line)
                orders_updated += 1
                
            elif not existing_order:
                # Crea nuovo ordine
                new_order = models.Order(
                    order_number=str(order_number),
                    customer_name=order_data["customer_name"]
                )
                db.add(new_order)
                db.flush()  # Per ottenere l'ID
                
                # Crea le righe d'ordine
                for sku, quantity in order_data["lines"].items():
                    order_line = models.OrderLine(
                        order_id=new_order.id,
                        product_sku=sku,
                        requested_quantity=quantity
                    )
                    db.add(order_line)
                orders_created += 1
            
            # LOGGING: Registra ogni prodotto dell'ordine corrente
            for sku, quantity in order_data["lines"].items():
                operation_type = OperationType.ORDINE_CREATO if is_new_order else OperationType.ORDINE_MODIFICATO
                
                logger.log_operation(
                    operation_type=operation_type,
                    operation_category=OperationCategory.FILE,
                    status=OperationStatus.SUCCESS,
                    product_sku=sku,
                    quantity=quantity,
                    user_id="excel_import_user",
                    file_name=f"ORDER_{order_number}",
                    details={
                        'order_number': str(order_number),
                        'customer_name': order_data["customer_name"],
                        'creation_method': 'excel_import_direct',
                        'source_file': file.filename,
                        'operation_description': f"Import Excel diretto: {operation_type.lower()} ordine {order_number}, aggiunto {quantity}x {sku} per cliente {order_data['customer_name']}",
                        'import_stats': {
                            'orders_created': orders_created,
                            'orders_updated': orders_updated,
                            'source_filename': file.filename
                        }
                    },
                    api_endpoint="/orders/parse-excel-orders"
                )
        
        db.commit()
        
        # Risposta con risultato finale
        result = {
            "success": True,
            "file_name": file.filename,
            "orders_created": orders_created,
            "orders_updated": orders_updated,
            "summary": {
                "total_orders": total_orders,
                "total_lines": total_lines,
                "errors": errors_count,
                "warnings": warnings_count,
                "orders_preview": list(orders_data.keys())[:5]
            },
            "message": f"Import Excel completato: {orders_created} ordini creati, {orders_updated} ordini aggiornati. {errors_count} errori saltati."
        }
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante parsing Excel: {str(e)}")

@router.post("/commit-excel-orders")
async def commit_excel_orders(
    request_data: dict,
    db: Session = Depends(get_db)
):
    """
    Commit degli ordini Excel dopo validazione e conferma del recap.
    Riceve i dati modificati dall'utente tramite il recap overlay.
    """
    try:
        recap_items = request_data.get("recap_items", [])
        file_name = request_data.get("file_name", "excel_import.xlsx")
        
        if not recap_items:
            raise HTTPException(status_code=400, detail="Nessun dato da importare")
        
        # Riorganizza i dati per ordine
        orders_to_create = defaultdict(lambda: {"customer_name": "", "lines": []})
        
        for item in recap_items:
            if item.get("status") == "ok":  # Elabora solo elementi validi
                order_number = item["order_number"]
                orders_to_create[order_number]["customer_name"] = item["customer_name"]
                orders_to_create[order_number]["lines"].append({
                    "product_sku": item["sku"],
                    "requested_quantity": item["quantity"]
                })
        
        if not orders_to_create:
            raise HTTPException(status_code=400, detail="Nessun ordine valido da creare")
        
        # Crea gli ordini nel database con logging integrato
        orders_created = 0
        orders_updated = 0
        logger = LoggingService(db)
        
        for order_number, order_data in orders_to_create.items():
            # Controlla se l'ordine esiste già
            existing_order = db.query(models.Order).filter(models.Order.order_number == order_number).first()
            is_new_order = existing_order is None
            
            if existing_order and not existing_order.is_completed:
                # Aggiorna ordine esistente
                for line_data in order_data["lines"]:
                    # Controlla se la riga prodotto esiste già
                    existing_line = db.query(models.OrderLine).filter(
                        models.OrderLine.order_id == existing_order.id,
                        models.OrderLine.product_sku == line_data["product_sku"]
                    ).first()
                    
                    if existing_line:
                        existing_line.requested_quantity += line_data["requested_quantity"]
                    else:
                        new_line = models.OrderLine(
                            order_id=existing_order.id,
                            product_sku=line_data["product_sku"],
                            requested_quantity=line_data["requested_quantity"]
                        )
                        db.add(new_line)
                orders_updated += 1
                
            elif not existing_order:
                # Crea nuovo ordine
                new_order = models.Order(
                    order_number=order_number,
                    customer_name=order_data["customer_name"]
                )
                db.add(new_order)
                db.flush()  # Per ottenere l'ID
                
                # Crea le righe d'ordine
                for line_data in order_data["lines"]:
                    order_line = models.OrderLine(
                        order_id=new_order.id,
                        product_sku=line_data["product_sku"],
                        requested_quantity=line_data["requested_quantity"]
                    )
                    db.add(order_line)
                orders_created += 1
            
            # LOGGING: Registra ogni prodotto dell'ordine corrente
            for line_data in order_data["lines"]:
                operation_type = OperationType.ORDINE_CREATO if is_new_order else OperationType.ORDINE_MODIFICATO
                
                logger.log_operation(
                    operation_type=operation_type,
                    operation_category=OperationCategory.FILE,
                    status=OperationStatus.SUCCESS,
                    product_sku=line_data["product_sku"],
                    quantity=line_data["requested_quantity"],
                    user_id="excel_import_user",
                    file_name=f"ORDER_{order_number}",
                    details={
                        'order_number': order_number,
                        'customer_name': order_data["customer_name"],
                        'creation_method': 'excel_import',
                        'source_file': file_name,
                        'operation_description': f"Import Excel: {operation_type.value.lower()} ordine {order_number}, aggiunto {line_data['requested_quantity']}x {line_data['product_sku']} per cliente {order_data['customer_name']}",
                        'import_stats': {
                            'orders_created': orders_created,
                            'orders_updated': orders_updated,
                            'source_filename': file_name
                        }
                    },
                    api_endpoint="/orders/commit-excel-orders"
                )
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Import Excel completato: {orders_created} ordini creati, {orders_updated} ordini aggiornati",
            "orders_created": orders_created,
            "orders_updated": orders_updated,
            "file_name": file_name
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Errore durante commit Excel: {str(e)}")

# === ENDPOINT MODIFICA NOME CLIENTE ===

from pydantic import BaseModel

class CustomerNameUpdateRequest(BaseModel):
    customer_name: str
    update_ddt: bool = False

@router.patch("/{order_number}/customer")
def update_customer_name(
    order_number: str, 
    update_request: CustomerNameUpdateRequest,
    db: Session = Depends(get_db)
):
    """Aggiorna il nome cliente di un ordine e opzionalmente del DDT collegato"""
    logger = LoggingService(db)
    
    try:
        # Trova l'ordine
        order = db.query(models.Order).filter(
            models.Order.order_number == order_number
        ).first()
        
        if not order:
            raise HTTPException(status_code=404, detail="Ordine non trovato")
        
        # Verifica che l'ordine non sia archiviato
        if order.is_archived:
            raise HTTPException(status_code=400, detail="Impossibile modificare ordini archiviati")
        
        old_customer_name = order.customer_name
        new_customer_name = update_request.customer_name.strip()
        
        if not new_customer_name:
            raise HTTPException(status_code=400, detail="Il nome cliente non può essere vuoto")
        
        if old_customer_name == new_customer_name:
            return {"success": True, "message": "Nessuna modifica necessaria"}
        
        # Aggiorna il nome cliente dell'ordine
        order.customer_name = new_customer_name
        
        updated_components = ["ordine"]
        
        # Se richiesto, aggiorna anche il DDT collegato
        if update_request.update_ddt:
            from wms_app.models.ddt import DDT
            
            ddt = db.query(DDT).filter(DDT.order_number == order_number).first()
            if ddt:
                ddt.customer_name = new_customer_name
                updated_components.append("DDT")
        
        # Log dell'operazione
        logger.log_operation(
            operation_type=OperationType.ORDINE_MODIFICATO,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.SUCCESS,
            product_sku=None,
            location_from=None,
            location_to=None,
            quantity=None,
            user_id="system",
            file_name=None,
            details={
                'order_number': order_number,
                'old_customer_name': old_customer_name,
                'new_customer_name': new_customer_name,
                'updated_components': updated_components,
                'operation_description': f"Modifica nome cliente ordine {order_number}: '{old_customer_name}' → '{new_customer_name}'"
            },
            api_endpoint=f"/orders/{order_number}/customer"
        )
        
        db.commit()
        
        components_text = " e ".join(updated_components)
        return {
            "success": True,
            "message": f"Nome cliente aggiornato con successo per {components_text}",
            "old_customer_name": old_customer_name,
            "new_customer_name": new_customer_name,
            "updated_components": updated_components
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.log_operation(
            operation_type=OperationType.ORDINE_MODIFICATO,
            operation_category=OperationCategory.MANUAL,
            status=OperationStatus.ERROR,
            error_message=str(e),
            details={
                'order_number': order_number,
                'attempted_customer_name': update_request.customer_name,
                'operation_description': f"Errore modifica nome cliente ordine {order_number}"
            },
            api_endpoint=f"/orders/{order_number}/customer"
        )
        raise HTTPException(status_code=500, detail=f"Errore durante l'aggiornamento: {str(e)}")